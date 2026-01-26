"""
Utility to clean Excel files - calculate formulas and remove images
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import pandas as pd
import sys


def clean_excel_pandas(input_file, output_file):
    """
    Clean Excel using pandas (for corrupted files)
    Note: This method loses formatting but can read corrupted files
    """
    # Read all sheets - try different engines
    all_sheets = None
    
    # Try openpyxl first
    try:
        all_sheets = pd.read_excel(input_file, sheet_name=None, engine='openpyxl')
    except Exception as e1:
        print(f"Warning: Could not read with openpyxl: {e1}")
        # Try xlrd for old .xls files
        try:
            all_sheets = pd.read_excel(input_file, sheet_name=None, engine='xlrd')
        except Exception as e2:
            print(f"Warning: Could not read with xlrd: {e2}")
            # Try calamine (if available) or default engine
            try:
                all_sheets = pd.read_excel(input_file, sheet_name=None)
            except Exception as e3:
                raise RuntimeError(f"Could not read file with any engine:\n- openpyxl: {e1}\n- xlrd: {e2}\n- default: {e3}")
    
    if all_sheets is None:
        raise RuntimeError("Failed to read file with pandas")

    # Filter sheets with "Tally" but not "Deck"
    filtered_sheets = {}
    for sheet_name, df in all_sheets.items():
        has_tally = 'Tally' in sheet_name or 'tally' in sheet_name
        has_deck = 'Deck' in sheet_name or 'deck' in sheet_name

        if has_tally and not has_deck:
            print(f"Keeping sheet: {sheet_name}")
            filtered_sheets[sheet_name] = df
        else:
            print(f"Skipping sheet: {sheet_name}")

    # Save to new file
    if filtered_sheets:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in filtered_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        print(f"Done! File saved: {output_file}")
    else:
        raise ValueError("No tally sheets found in the file. Make sure the file has sheets with 'Tally' in the name (but not 'Deck').")


def clean_excel(input_file, output_file):
    """
    Clean Excel file by:
    1. Calculating all formulas and replacing them with values
    2. Removing all images
    3. Unmerging cells and filling with values
    4. Keeping only sheets with 'Tally' in name (but not 'Deck')
    5. Preserving cell formatting
    """
    print(f"Loading {input_file}...")

    # Try to load with openpyxl first
    wb_data = None
    wb_format = None
    wb_new = None

    try:
        # Load with data_only=True to get calculated values
        wb_data = load_workbook(input_file, data_only=True)
        # Load with formatting
        wb_format = load_workbook(input_file)
        use_openpyxl = True
    except Exception as e:
        print(f"Warning: Could not load with openpyxl ({e})")
        print("Trying with pandas (formatting will be lost)...")
        use_openpyxl = False

        # Clean up any partially loaded workbooks before falling back to pandas
        if wb_data:
            try:
                wb_data.close()
            except Exception:
                pass
            wb_data = None

        if wb_format:
            try:
                wb_format.close()
            except Exception:
                pass
            wb_format = None

    if not use_openpyxl:
        # Use pandas as fallback
        return clean_excel_pandas(input_file, output_file)

    try:
        # Create new workbook
        wb_new = Workbook()
        wb_new.remove(wb_new.active)  # Remove default sheet

        # Process each sheet
        for sheet_name in wb_data.sheetnames:
            # Filter: keep only sheets with "Tally" but not "Deck"
            has_tally = 'Tally' in sheet_name or 'tally' in sheet_name
            has_deck = 'Deck' in sheet_name or 'deck' in sheet_name

            if not (has_tally and not has_deck):
                print(f"Skipping sheet: {sheet_name} (not a tally sheet)")
                continue
            print(f"Processing sheet: {sheet_name}")

            ws_data = wb_data[sheet_name]
            ws_format = wb_format[sheet_name]
            ws_new = wb_new.create_sheet(title=sheet_name)

            # First, collect merged cell ranges and their values
            # Only fill the TOP-LEFT cell of merged range, leave others empty
            merged_cells_map = {}
            skip_cells = set()
            for merged_range in ws_format.merged_cells.ranges:
                # Get the top-left cell value
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                top_left_value = ws_data.cell(row=min_row, column=min_col).value

                # Store value only for top-left cell
                merged_cells_map[(min_row, min_col)] = top_left_value

                # Mark other cells in the range as empty
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        if row != min_row or col != min_col:
                            skip_cells.add((row, col))

            # Copy values and formatting
            for row_idx, row in enumerate(ws_data.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    # Skip cells that were part of merged range (except top-left)
                    if (row_idx, col_idx) in skip_cells:
                        value = None
                    # Get value from merged cells map or original cell
                    elif (row_idx, col_idx) in merged_cells_map:
                        value = merged_cells_map[(row_idx, col_idx)]
                    else:
                        value = cell.value

                    # Get formatting from original
                    orig_cell = ws_format.cell(row=row_idx, column=col_idx)

                    # Set value in new workbook
                    new_cell = ws_new.cell(row=row_idx, column=col_idx, value=value)

                    # Copy formatting
                    if orig_cell.has_style:
                        try:
                            new_cell.font = copy(orig_cell.font)
                            new_cell.border = copy(orig_cell.border)
                            new_cell.fill = copy(orig_cell.fill)
                            new_cell.number_format = orig_cell.number_format
                            new_cell.protection = copy(orig_cell.protection)
                            new_cell.alignment = copy(orig_cell.alignment)
                        except Exception as e:
                            pass  # Skip if formatting copy fails

            # Copy column widths
            for col in ws_format.column_dimensions:
                if col in ws_format.column_dimensions:
                    ws_new.column_dimensions[col].width = ws_format.column_dimensions[col].width

            # Copy row heights
            for row in ws_format.row_dimensions:
                if row in ws_format.row_dimensions:
                    ws_new.row_dimensions[row].height = ws_format.row_dimensions[row].height

            # Note: We don't copy merged cells as we've already unmerged them

        # Check if we have any sheets
        if len(wb_new.sheetnames) == 0:
            raise ValueError("No tally sheets found in the file. Make sure the file has sheets with 'Tally' in the name (but not 'Deck').")

        # Save cleaned file
        print(f"Saving to {output_file}...")
        wb_new.save(output_file)
        print(f"Done! File saved: {output_file}")

    except Exception as e:
        # If error occurs, try pandas fallback
        print(f"Error during processing: {e}")
        print("Falling back to pandas method...")
        return clean_excel_pandas(input_file, output_file)
    finally:
        # Close all workbooks to free resources
        if wb_data:
            try:
                wb_data.close()
            except Exception:
                pass

        if wb_format:
            try:
                wb_format.close()
            except Exception:
                pass

        if wb_new:
            try:
                wb_new.close()
            except Exception:
                pass


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python clean_excel.py <input_file> [output_file]")
        print("If output_file is not specified, will add '_cleaned' to input filename")
        sys.exit(1)

    input_file = sys.argv[1]

    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # Add _cleaned before extension
        parts = input_file.rsplit('.', 1)
        output_file = f"{parts[0]}_cleaned.{parts[1]}"

    clean_excel(input_file, output_file)
