#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Tally Converter - Graphical interface for processing tally files
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import subprocess
import platform
from pathlib import Path
import tempfile

# Import our modules
from utils.clean_excel import clean_excel
from tally_parser import parse_tally_file
from excel_generator import generate_excel
from manual_column_selector import ManualColumnSelector


class TallyParserGUI:
    """Main GUI application for Tally Parser"""

    def __init__(self, root):
        self.root = root
        self.root.title("Tally Parser")
        self.root.geometry("500x400")

        # Current file
        self.current_file = None
        self.cleaned_file = None
        self.converted_file = None  # Track converted .xls files for cleanup

        # Setup UI
        self._setup_ui()

    def _setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_label = ttk.Label(
            main_frame,
            text="Tally Parser",
            font=('Arial', 16, 'bold')
        )
        title_label.pack(pady=(0, 20))

        # Current file display
        self.file_label = ttk.Label(
            main_frame,
            text="No file selected",
            foreground="gray"
        )
        self.file_label.pack(pady=(0, 20))

        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.BOTH, expand=True)

        # Button 1: Open Tally File
        self.open_button = ttk.Button(
            button_frame,
            text="1. Open Tally File",
            command=self.open_file,
            width=30
        )
        self.open_button.pack(pady=10)

        # Button 2: Generate Clean Tally
        self.clean_button = ttk.Button(
            button_frame,
            text="2. Generate Clean Tally",
            command=self.generate_clean_tally,
            state=tk.DISABLED,
            width=30
        )
        self.clean_button.pack(pady=10)

        # Button 3: Generate TLY
        self.generate_button = ttk.Button(
            button_frame,
            text="3. Generate TLY",
            command=self.generate_tly,
            state=tk.DISABLED,
            width=30
        )
        self.generate_button.pack(pady=10)

        # Button Auto: Full automatic processing
        self.auto_button = ttk.Button(
            button_frame,
            text="Auto (Output + TLY)",
            command=self.auto_process,
            state=tk.DISABLED,
            width=30
        )
        self.auto_button.pack(pady=10)

        # Button 4: Readme
        self.readme_button = ttk.Button(
            button_frame,
            text="Readme",
            command=self.show_readme,
            width=30
        )
        self.readme_button.pack(pady=10)

        # Status bar
        self.status_label = ttk.Label(
            main_frame,
            text="Ready",
            relief=tk.SUNKEN,
            anchor=tk.W
        )
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X, pady=(20, 0))

    def _open_file_in_excel(self, file_path):
        """Open file in Excel (cross-platform)"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                return
            
            system = platform.system()
            
            if system == "Windows":
                # Windows: use os.startfile or subprocess
                try:
                    os.startfile(str(file_path))
                except Exception:
                    # Fallback: try to open with excel.exe (without console window)
                    creation_flags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
                    subprocess.Popen(['excel.exe', str(file_path)], shell=False, creationflags=creation_flags)
            elif system == "Darwin":  # macOS
                subprocess.Popen(['open', '-a', 'Microsoft Excel', str(file_path)], 
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:  # Linux and others
                # Try common Linux Excel alternatives
                subprocess.Popen(['xdg-open', str(file_path)], 
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            # Silently fail - don't interrupt the workflow if Excel can't be opened
            pass

    def _open_file_in_notepad(self, file_path):
        """Open file in Notepad (Windows) or default text editor"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                return
            
            system = platform.system()
            
            if system == "Windows":
                # Windows: use notepad (without console window)
                try:
                    creation_flags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
                    subprocess.Popen(['notepad.exe', str(file_path)], 
                                   shell=False, creationflags=creation_flags)
                except Exception:
                    # Fallback: use os.startfile
                    os.startfile(str(file_path))
            elif system == "Darwin":  # macOS
                subprocess.Popen(['open', '-a', 'TextEdit', str(file_path)],
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:  # Linux and others
                # Try common Linux text editors
                subprocess.Popen(['xdg-open', str(file_path)],
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            # Silently fail - don't interrupt the workflow if editor can't be opened
            pass

    def _convert_xls_to_xlsx(self, xls_file):
        """Convert .xls file to .xlsx format"""
        try:
            import pandas as pd
            
            self.status_label.config(text="Converting .xls to .xlsx...")
            self.root.update()
            
            # Read .xls file using pandas with xlrd engine
            try:
                # Try with xlrd engine (for .xls files)
                all_sheets = pd.read_excel(xls_file, sheet_name=None, engine='xlrd')
            except ImportError:
                # If xlrd is not installed, try with openpyxl (might work for some files)
                try:
                    all_sheets = pd.read_excel(xls_file, sheet_name=None, engine='openpyxl')
                except Exception:
                    raise ImportError("xlrd package is required to read .xls files. Install it with: pip install xlrd")
            except Exception as e:
                # Other errors reading the file
                raise Exception(f"Failed to read .xls file: {str(e)}")
            
            # Create .xlsx file in same directory as original
            xls_path = Path(xls_file)
            xlsx_file = xls_path.parent / f"{xls_path.stem}_converted.xlsx"
            
            # Delete existing converted file if it exists
            if xlsx_file.exists():
                try:
                    xlsx_file.unlink()
                except Exception:
                    pass
            
            # Write to .xlsx using openpyxl engine
            with pd.ExcelWriter(str(xlsx_file), engine='openpyxl') as writer:
                for sheet_name, df in all_sheets.items():
                    # Truncate sheet name if too long (Excel limit is 31 characters)
                    sheet_name_truncated = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                    df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)
            
            self.status_label.config(text=f"Converted: {xlsx_file.name}")
            self.root.update()
            
            messagebox.showinfo(
                "File Converted",
                f"Old .xls file has been converted to .xlsx format.\n\n"
                f"Original: {xls_path.name}\n"
                f"Converted: {xlsx_file.name}\n\n"
                f"You can now proceed with parsing."
            )
            
            return str(xlsx_file)
            
        except ImportError as e:
            messagebox.showerror(
                "Missing Package",
                f"xlrd package is required to convert .xls files.\n\n"
                f"Please install it:\npip install xlrd\n\n"
                f"Error: {str(e)}"
            )
            return None
        except Exception as e:
            messagebox.showerror("Conversion Error", f"Failed to convert .xls to .xlsx:\n{str(e)}")
            self.status_label.config(text="Conversion failed")
            return None

    def open_file(self):
        """Open a tally file"""
        filename = filedialog.askopenfilename(
            title="Select Tally File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xlsm"),
                ("All files", "*.*")
            ]
        )

        if filename:
            file_path = Path(filename)
            
            # Delete existing related files to avoid conflicts
            file_stem = file_path.stem
            file_dir = file_path.parent
            
            related_files = [
                file_dir / f"{file_stem}_converted.xlsx",
                file_dir / f"{file_stem}_cleaned.xlsx",
                file_dir / f"{file_stem}_output.xlsx"
            ]
            
            for related_file in related_files:
                if related_file.exists():
                    try:
                        related_file.unlink()
                    except Exception:
                        pass
            
            # Check if it's an old .xls file
            if file_path.suffix.lower() == '.xls':
                # Convert to .xlsx first
                converted_file = self._convert_xls_to_xlsx(filename)
                if not converted_file:
                    # Conversion failed, don't proceed
                    return
                # Use converted file
                self.current_file = converted_file
                self.converted_file = converted_file  # Track for cleanup
                original_name = file_path.name
                converted_name = Path(converted_file).name
            else:
                # Use file as-is
                self.current_file = filename
                self.converted_file = None
                original_name = file_path.name
                converted_name = None
            
            self.cleaned_file = None

            # Update UI
            if converted_name:
                self.file_label.config(
                    text=f"Selected: {original_name} (converted to {converted_name})",
                    foreground="black"
                )
            else:
                self.file_label.config(
                    text=f"Selected: {original_name}",
                    foreground="black"
                )

            # Enable buttons
            self.clean_button.config(state=tk.NORMAL)
            # Generate TLY button can work independently (reads from any xlsx file)
            self.generate_button.config(state=tk.NORMAL)
            # Auto button enabled when file is loaded
            self.auto_button.config(state=tk.NORMAL)

            if converted_name:
                self.status_label.config(text=f"Loaded: {original_name} → {converted_name}")
            else:
                self.status_label.config(text=f"Loaded: {original_name}")

    def generate_clean_tally(self):
        """Generate clean tally - show Auto/Choose dialog, then automatically generate output"""
        if not self.current_file:
            messagebox.showerror("Error", "Please open a file first")
            return

        input_path = Path(self.current_file)
        
        # If file is .xls, convert to .xlsx first using pandas + xlrd
        if input_path.suffix.lower() == '.xls':
            converted_file = self._convert_xls_to_xlsx(str(input_path))
            if not converted_file:
                return
            self.current_file = converted_file
            self.converted_file = converted_file
            input_path = Path(self.current_file)

        # Delete existing cleaned and output files before generation
        cleaned_file_path = input_path.parent / f"{input_path.stem}_cleaned.xlsx"
        output_file_path = input_path.parent / f"{input_path.stem}_output.xlsx"

        if cleaned_file_path.exists():
            try:
                os.remove(cleaned_file_path)
            except Exception:
                pass

        if output_file_path.exists():
            try:
                os.remove(output_file_path)
            except Exception:
                pass

        # Show dialog: Auto or Choose column header
        dialog = ModeSelectionDialog(self.root, self.current_file, self.status_label)
        self.root.wait_window(dialog.dialog)

        # Check result
        if dialog.result:
            mode, cleaned_file = dialog.result

            if cleaned_file:
                self.cleaned_file = cleaned_file
                
                # Step 1: Clean file is ready
                self.status_label.config(text=f"Cleaned file ready: {os.path.basename(cleaned_file)}")
                self.root.update()
                
                # Step 2: Automatically parse and generate output
                try:
                    # Ask for output location
                    output_file = filedialog.asksaveasfilename(
                        title="Save TLY Output",
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                        initialfile=Path(self.current_file).stem + "_output.xlsx"
                    )

                    if not output_file:
                        # User cancelled - just enable the button for manual generation
                        self.generate_button.config(state=tk.NORMAL)
                        self.status_label.config(text=f"Cleaned file ready: {os.path.basename(cleaned_file)}")
                        messagebox.showinfo(
                            "Success",
                            f"Clean tally generated:\n{os.path.basename(cleaned_file)}\n\nYou can generate TLY output manually using button 3."
                        )
                        return

                    # Parse data
                    self.status_label.config(text="Parsing data...")
                    self.root.update()

                    data = parse_tally_file(self.cleaned_file)

                    if not data:
                        messagebox.showwarning("Warning", "No data extracted from file")
                        self.generate_button.config(state=tk.NORMAL)
                        return

                    # Generate output
                    self.status_label.config(text="Generating output file...")
                    self.root.update()

                    generate_excel(data, output_file)

                    # Success
                    self.generate_button.config(state=tk.NORMAL)
                    self.status_label.config(text=f"Generated: {os.path.basename(output_file)}")
                    
                    # Open file in Excel
                    self._open_file_in_excel(output_file)
                    
                    messagebox.showinfo(
                        "Success",
                        f"Clean tally and TLY output generated successfully!\n\n"
                        f"Cleaned file: {os.path.basename(cleaned_file)}\n"
                        f"Output file: {os.path.basename(output_file)}\n"
                        f"Rows processed: {len(data)}\n\n"
                        f"File opened in Excel."
                    )

                except Exception as e:
                    error_msg = str(e)
                    # Check if it's a header detection error
                    if "Could not find header row" in error_msg or "required columns" in error_msg:
                        detailed_msg = (
                            f"Не удалось автоматически найти заголовки таблицы.\n\n"
                            f"Возможные причины:\n"
                            f"• Заголовки находятся не в первых 30 строках\n"
                            f"• Названия колонок отличаются от ожидаемых\n"
                            f"• Структура таблицы нестандартная\n\n"
                            f"Рекомендация: Используйте режим 'Choose Column Header'\n"
                            f"для ручного выбора заголовков колонок."
                        )
                        messagebox.showerror("Ошибка поиска заголовков", detailed_msg)
                    else:
                        messagebox.showerror("Error", f"Failed to generate output:\n{error_msg}")
                    self.status_label.config(text="Error occurred")
                    self.generate_button.config(state=tk.NORMAL)

    def auto_process(self):
        """Auto process: Generate cleaned file, output.xlsx, and TLY file automatically"""
        if not self.current_file:
            messagebox.showerror("Error", "Please open a file first")
            return

        input_path = Path(self.current_file)
        
        # If file is .xls, convert to .xlsx first
        if input_path.suffix.lower() == '.xls':
            converted_file = self._convert_xls_to_xlsx(str(input_path))
            if not converted_file:
                return
            self.current_file = converted_file
            self.converted_file = converted_file
            input_path = Path(self.current_file)

        # Delete existing files before generation
        cleaned_file_path = input_path.parent / f"{input_path.stem}_cleaned.xlsx"
        output_file_path = input_path.parent / f"{input_path.stem}_output.xlsx"
        tly_file_path = input_path.parent / f"{input_path.stem}_output.tly"

        for file_path in [cleaned_file_path, output_file_path, tly_file_path]:
            if file_path.exists():
                try:
                    os.remove(file_path)
                except Exception:
                    pass

        try:
            # Step 1: Generate cleaned file using Auto mode
            self.status_label.config(text="Generating cleaned file (Auto mode)...")
            self.root.update()

            # Use Auto mode directly (bypass dialog)
            from utils.clean_excel import clean_excel
            cleaned_file = str(cleaned_file_path)
            
            try:
                clean_excel(str(input_path), cleaned_file)
                
                # Verify that cleaned file was created
                if not Path(cleaned_file).exists():
                    raise FileNotFoundError(f"Cleaned file was not created: {cleaned_file}")
            except Exception as e:
                # If cleaning failed, try fallback: calculate formulas in Excel and retry
                try:
                    from utils.calculate_formulas import calculate_formulas
                    temp_calculated = input_path.parent / f"{input_path.stem}_calculated_temp.xlsx"
                    
                    self.status_label.config(text="Opening file in Excel...")
                    self.root.update()
                    
                    calculate_formulas(str(input_path), str(temp_calculated))
                    
                    if not temp_calculated.exists():
                        raise FileNotFoundError(f"Calculated file was not created: {temp_calculated}")
                    
                    self.status_label.config(text="Processing calculated file...")
                    self.root.update()
                    
                    try:
                        clean_excel(str(temp_calculated), cleaned_file)
                    except Exception as clean_error:
                        from utils.clean_excel import clean_excel_pandas
                        clean_excel_pandas(str(temp_calculated), cleaned_file)
                    
                    if not Path(cleaned_file).exists():
                        raise FileNotFoundError(f"Cleaned file was not created: {cleaned_file}")
                    
                    # Clean up temporary file
                    if temp_calculated.exists():
                        try:
                            temp_calculated.unlink()
                        except Exception:
                            pass
                except Exception as e2:
                    raise Exception(f"Failed to clean file: {str(e)}\nFallback also failed: {str(e2)}")

            # Step 2: Parse and generate output.xlsx
            self.status_label.config(text="Parsing data...")
            self.root.update()

            data = parse_tally_file(cleaned_file)

            if not data:
                messagebox.showwarning("Warning", "No data extracted from file")
                return

            self.status_label.config(text="Generating output file...")
            self.root.update()

            output_file = str(output_file_path)
            generate_excel(data, output_file)

            # Step 3: Generate TLY file
            self.status_label.config(text="Generating TLY file...")
            self.root.update()

            from openpyxl import load_workbook

            wb = None
            try:
                wb = load_workbook(output_file, data_only=True)
                ws = wb.active

                # Find column indices for "Item #" and "Depth"
                header_row = 1
                item_col = None
                depth_col = None

                for col_idx, cell in enumerate(ws[header_row], 1):
                    cell_value = str(cell.value or "").strip()
                    if "item" in cell_value.lower() and "#" in cell_value:
                        item_col = col_idx
                    elif "depth" in cell_value.lower():
                        depth_col = col_idx

                if item_col is None or depth_col is None:
                    messagebox.showerror("Error", "Could not find 'Item #' and 'Depth' columns in the output file")
                    return

                # Extract data for TLY
                rows_data = []
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    item_value = ws.cell(row=row_idx, column=item_col).value
                    depth_value = ws.cell(row=row_idx, column=depth_col).value

                    if item_value is None and depth_value is None:
                        continue

                    item_str = str(item_value).strip() if item_value is not None else ""
                    depth_str = str(depth_value).strip() if depth_value is not None else ""

                    if depth_str:
                        rows_data.append((item_str, depth_str))

                if not rows_data:
                    messagebox.showwarning("Warning", "No data found for TLY file")
                    return

            finally:
                # Close workbook to free resources
                if wb:
                    try:
                        wb.close()
                    except Exception:
                        pass

            # Generate TLY file
            tly_file = str(tly_file_path)
            with open(tly_file, 'w', encoding='utf-8') as f:
                f.write("Run Number\tDepth of Top of Joint\n")
                for item, depth in rows_data:
                    f.write(f"{item}\t{depth}\n")

            # Success - open files
            self.status_label.config(text=f"Success! Generated output and TLY files")
            self.root.update()

            # Open Excel with output file
            self._open_file_in_excel(output_file)
            
            # Open Notepad with TLY file
            self._open_file_in_notepad(tly_file)
            
            messagebox.showinfo(
                "Success",
                f"Auto processing completed successfully!\n\n"
                f"Files generated:\n"
                f"• Output: {os.path.basename(output_file)}\n"
                f"• TLY: {os.path.basename(tly_file)}\n"
                f"Rows processed: {len(data)}\n\n"
                f"Files opened in Excel and Notepad."
            )

        except Exception as e:
            error_msg = str(e)
            if "Could not find header row" in error_msg or "required columns" in error_msg:
                detailed_msg = (
                    f"Не удалось автоматически найти заголовки таблицы.\n\n"
                    f"Возможные причины:\n"
                    f"• Заголовки находятся не в первых 50 строках\n"
                    f"• Названия колонок отличаются от ожидаемых\n"
                    f"• Структура таблицы нестандартная\n\n"
                    f"Рекомендация: Используйте кнопку '2. Generate Clean Tally'\n"
                    f"с режимом 'Choose Column Header' для ручного выбора."
                )
                messagebox.showerror("Ошибка поиска заголовков", detailed_msg)
            else:
                messagebox.showerror("Error", f"Auto processing failed:\n{error_msg}")
            self.status_label.config(text="Error occurred")

    def generate_tly(self):
        """Generate final TLY output from output xlsx file"""
        try:
            # Ask user to select output xlsx file
            input_xlsx = filedialog.askopenfilename(
                title="Select Output XLSX File",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialdir=Path(self.current_file).parent if self.current_file else None
            )

            if not input_xlsx:
                return

            # Ask for TLY output location
            tly_file = filedialog.asksaveasfilename(
                title="Save TLY File",
                defaultextension=".tly",
                filetypes=[("TLY files", "*.tly"), ("Text files", "*.txt"), ("All files", "*.*")],
                initialfile=Path(input_xlsx).stem + ".tly"
            )

            if not tly_file:
                return

            # Read Excel file and extract Item # and Depth columns
            self.status_label.config(text="Reading Excel file...")
            self.root.update()

            from openpyxl import load_workbook

            wb = None
            try:
                wb = load_workbook(input_xlsx, data_only=True)
                ws = wb.active

                # Find column indices for "Item #" and "Depth"
                header_row = 1
                item_col = None
                depth_col = None

                for col_idx, cell in enumerate(ws[header_row], 1):
                    cell_value = str(cell.value or "").strip()
                    if "item" in cell_value.lower() and "#" in cell_value:
                        item_col = col_idx
                    elif "depth" in cell_value.lower():
                        depth_col = col_idx

                if item_col is None or depth_col is None:
                    messagebox.showerror("Error", "Could not find 'Item #' and 'Depth' columns in the Excel file")
                    return

                # Extract data
                self.status_label.config(text="Extracting data...")
                self.root.update()

                rows_data = []
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    item_value = ws.cell(row=row_idx, column=item_col).value
                    depth_value = ws.cell(row=row_idx, column=depth_col).value

                    # Skip empty rows
                    if item_value is None and depth_value is None:
                        continue

                    # Format values
                    item_str = str(item_value).strip() if item_value is not None else ""
                    depth_str = str(depth_value).strip() if depth_value is not None else ""

                    if depth_str:  # Only add rows with depth
                        rows_data.append((item_str, depth_str))

                if not rows_data:
                    messagebox.showwarning("Warning", "No data found in the Excel file")
                    return

            finally:
                # Close workbook to free resources
                if wb:
                    try:
                        wb.close()
                    except Exception:
                        pass

            # Generate TLY file
            self.status_label.config(text="Generating TLY file...")
            self.root.update()

            with open(tly_file, 'w', encoding='utf-8') as f:
                # Write header
                f.write("Run Number\tDepth of Top of Joint\n")

                # Write data rows
                for item, depth in rows_data:
                    f.write(f"{item}\t{depth}\n")

            # Success
            self.status_label.config(text=f"Generated: {os.path.basename(tly_file)}")
            
            # Open file in Notepad
            self._open_file_in_notepad(tly_file)
            
            messagebox.showinfo(
                "Success",
                f"TLY file generated successfully!\n\nRows: {len(rows_data)}\nFile: {os.path.basename(tly_file)}\n\nFile opened in Notepad."
            )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate TLY:\n{str(e)}")
            self.status_label.config(text="Error occurred")
            import traceback
            traceback.print_exc()

    def show_readme(self):
        """Show readme window with instructions"""
        # Create readme window
        readme_window = tk.Toplevel(self.root)
        readme_window.title("Help - Tally Parser")
        readme_window.geometry("750x650")
        readme_window.transient(self.root)
        readme_window.resizable(True, True)
        
        # Center window
        readme_window.update_idletasks()
        x = (readme_window.winfo_screenwidth() // 2) - (750 // 2)
        y = (readme_window.winfo_screenheight() // 2) - (650 // 2)
        readme_window.geometry(f"750x650+{x}+{y}")
        
        # Create main container
        main_frame = ttk.Frame(readme_window, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text="Tally Parser User Guide",
            font=('Arial', 14, 'bold')
        )
        title_label.pack(pady=(0, 15))
        
        # Create scrollable text widget with proper layout
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Text widget with scrollbar
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            font=('Consolas', 9),
            padx=15,
            pady=15,
            bg='white',
            relief=tk.FLAT,
            borderwidth=1
        )
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        # Readme content
        readme_text = """TALLY PARSER USER GUIDE


1. OPENING A FILE

   • Click the "1. Open Tally File" button
   • Select an Excel file (.xls or .xlsx) with well completion data
   • The program will automatically detect the file format
   • If the file is in the old .xls format, it will be automatically 
     converted to .xlsx

   Note: The file must contain sheets with names containing "Tally" 
   (but not "Deck Tally").


2. GENERATING CLEAN TALLY

   Click the "2. Generate Clean Tally" button. A mode selection dialog 
   will appear:

   AUTOMATIC MODE (Auto)
   
   • The program will automatically detect table headers
   • Will find columns: Depth, Length, Item #, Comments
   • Will perform cleaning and generate output file
   • The output file will automatically open in Excel
   
   Use this mode if the file structure is standard

   MANUAL MODE (Choose Column Header)
   
   Step 1: Select Header Row
   • The program will show a table with data
   • Will automatically detect the header row
   • Click on the row with headers (Item, Depth, Length...)
   • Click "Confirm header row"

   Step 2: Select Depth Column
   • Click on the column header with depth
   • The column will be highlighted
   • Click "Confirm depth column"

   Step 3: Select Effective Length Column
   • Click on the column header with length
   • Click "Confirm length column"

   Step 4: Select Item Number Column
   • Click on the column header with item number
   • Click "Confirm item number column"

   Step 5: Select Comments Column
   • Click on the column header with comments
   • Click "Confirm comments column"

   After selecting all columns, processing will begin

   Use this mode if automatic detection doesn't work correctly

   Result:
   • Creates [filename]_cleaned.xlsx file (cleaned file)
   • Creates [filename]_output.xlsx file (output file)
   • The output file automatically opens in Excel


3. GENERATING TLY FILE

   • Click the "3. Generate TLY" button
   • Select a previously created output .xlsx file
   • The program will extract "Depth" and "Item #" columns
   • Will create a .tly text file in the format:
     [Run Number]    [Depth of Top of Joint]
   • The file will automatically open in Notepad

   Note: This function works independently and can be used for any 
   output .xlsx file, even if it was created earlier.


IMPORTANT NOTES

   • Before generation, existing _cleaned.xlsx and _output.xlsx files 
     are automatically deleted
   
   • Files must contain sheets with names including "Tally"
     (sheets named "Deck Tally" are ignored)
   
   • If a file is in "Page Layout" view, it will be automatically 
     switched to normal view
   
   • To work with .xls files, the xlrd package is required:
     pip install xlrd


STATUS BAR

   At the bottom of the window, the current program status is displayed:
   • "Ready" - program is ready to work
   • "Loaded: [filename]" - file is loaded
   • "Cleaned file ready: [filename]" - file is cleaned
   • "Processing..." - processing in progress
   • Error messages when they occur


WORKFLOW

   1. Open a file (button 1)
   2. Generate Clean Tally (button 2) - select mode (Auto/Manual)
   3. If needed, generate TLY file (button 3)

   All files are saved in the same folder as the source file.
"""
        
        text_widget.insert(tk.END, readme_text)
        text_widget.config(state=tk.DISABLED)  # Make read-only
        
        # Close button
        close_button = ttk.Button(
            main_frame,
            text="Close",
            command=readme_window.destroy
        )
        close_button.pack(pady=(10, 0))


class ModeSelectionDialog:
    """Dialog to choose Auto or Manual column selection mode"""

    def __init__(self, parent, file_path, status_label=None):
        self.file_path = file_path
        self.result = None
        self.status_label = status_label  # Reference to parent's status label

        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Mode")
        self.dialog.geometry("400x200")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Center dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (200 // 2)
        self.dialog.geometry(f"400x200+{x}+{y}")

        self._setup_ui()

    def _setup_ui(self):
        """Setup dialog UI"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_label = ttk.Label(
            main_frame,
            text="Choose Column Header Detection Mode",
            font=('Arial', 11, 'bold')
        )
        title_label.pack(pady=(0, 20))

        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(expand=True)

        # Auto button
        auto_button = ttk.Button(
            button_frame,
            text="Auto",
            command=self.use_auto_mode,
            width=15
        )
        auto_button.pack(side=tk.LEFT, padx=10)

        # Auto description
        auto_desc = ttk.Label(
            button_frame,
            text="Automatic detection",
            foreground="gray"
        )
        auto_desc.pack(side=tk.LEFT, padx=5)

        # Separator
        ttk.Separator(main_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=20)

        # Manual button frame
        manual_frame = ttk.Frame(main_frame)
        manual_frame.pack()

        # Choose button
        choose_button = ttk.Button(
            manual_frame,
            text="Choose Column Header",
            command=self.use_manual_mode,
            width=20
        )
        choose_button.pack(side=tk.LEFT, padx=10)

        # Manual description
        manual_desc = ttk.Label(
            manual_frame,
            text="Manual selection",
            foreground="gray"
        )
        manual_desc.pack(side=tk.LEFT, padx=5)

    def use_auto_mode(self):
        """Use automatic header detection"""
        try:
            input_path = Path(self.file_path)
            
            # If file is .xls, convert to .xlsx first using pandas + xlrd
            if input_path.suffix.lower() == '.xls':
                import pandas as pd
                try:
                    # Read .xls file using pandas with xlrd engine
                    all_sheets = pd.read_excel(self.file_path, sheet_name=None, engine='xlrd')
                    
                    # Create .xlsx file path
                    xlsx_file = input_path.parent / f"{input_path.stem}_converted.xlsx"
                    
                    # Delete existing converted file if it exists
                    if xlsx_file.exists():
                        try:
                            xlsx_file.unlink()
                        except Exception:
                            pass
                    
                    # Write to .xlsx using openpyxl engine
                    with pd.ExcelWriter(str(xlsx_file), engine='openpyxl') as writer:
                        for sheet_name, df in all_sheets.items():
                            # Truncate sheet name if too long (Excel limit is 31 characters)
                            sheet_name_truncated = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                            df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)
                    
                    # Use converted file for cleaning
                    input_path = xlsx_file
                    
                except ImportError as e:
                    messagebox.showerror(
                        "Missing Package",
                        f"xlrd package is required to convert .xls files.\n\n"
                        f"Please install it:\npip install xlrd\n\n"
                        f"Error: {str(e)}"
                    )
                    self.result = None
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to convert .xls to .xlsx:\n{str(e)}")
                    self.result = None
                    return
            
            # Use existing clean_excel function
            output_path = input_path.parent / f"{input_path.stem}_cleaned.xlsx"

            # Clean the file
            try:
                clean_excel(str(input_path), str(output_path))
                
                # Verify that cleaned file was created
                if not output_path.exists():
                    raise FileNotFoundError(f"Cleaned file was not created: {output_path}")
                
                self.result = ('auto', str(output_path))
                self.dialog.destroy()
            except Exception as e:
                # If cleaning failed, try fallback: calculate formulas in Excel and retry
                try:
                    # Ask user if they want to try fallback method
                    response = messagebox.askyesno(
                        "Error",
                        "Can't parse Excel book, try to fix it in Excel?"
                    )
                    
                    if response:
                        # Create temporary file for Excel calculation
                        temp_calculated = input_path.parent / f"{input_path.stem}_calculated_temp.xlsx"
                        
                        # Calculate formulas using Excel COM
                        try:
                            from utils.calculate_formulas import calculate_formulas
                            
                            # Show progress message
                            if self.status_label:
                                self.status_label.config(text="Opening file in Excel...")
                            if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                                self.dialog.update()
                            
                            # Calculate formulas - this will open Excel
                            calculate_formulas(str(input_path), str(temp_calculated))
                            
                            # Verify that calculated file was created
                            if not temp_calculated.exists():
                                raise FileNotFoundError(f"Calculated file was not created: {temp_calculated}")
                            
                            if self.status_label:
                                self.status_label.config(text="Processing calculated file...")
                            if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                                self.dialog.update()
                            
                            # Try cleaning again with calculated file
                            # If clean_excel fails, try using pandas fallback
                            try:
                                clean_excel(str(temp_calculated), str(output_path))
                            except Exception as clean_error:
                                # If clean_excel fails, try pandas fallback directly
                                from utils.clean_excel import clean_excel_pandas
                                clean_excel_pandas(str(temp_calculated), str(output_path))
                            
                            # Verify that cleaned file was created
                            if not output_path.exists():
                                raise FileNotFoundError(f"Cleaned file was not created: {output_path}")
                            
                            # Clean up temporary file
                            if temp_calculated.exists():
                                try:
                                    temp_calculated.unlink()
                                except Exception:
                                    pass
                            
                            self.result = ('auto', str(output_path))
                            self.dialog.destroy()
                        except ImportError as e_import:
                            messagebox.showerror(
                                "Missing Package",
                                f"Required package not found:\n{str(e_import)}\n\n"
                                "Please install pywin32:\npip install pywin32"
                            )
                            self.result = None
                            return
                        except Exception as e2:
                            # Clean up temporary file if it exists
                            if 'temp_calculated' in locals() and temp_calculated.exists():
                                try:
                                    temp_calculated.unlink()
                                except Exception:
                                    pass
                            
                            # Show detailed error
                            error_details = str(e2)
                            import traceback
                            traceback.print_exc()
                            
                            messagebox.showerror(
                                "Fallback Error",
                                f"Fallback method failed:\n{error_details}\n\n"
                                "Please try using manual column selection mode."
                            )
                            self.result = None
                            return
                    else:
                        self.result = None
                        return
                except Exception as e_fallback:
                    # If fallback also failed, show error
                    messagebox.showerror(
                        "Error",
                        f"Failed to clean file:\n{str(e)}\n\n"
                        f"Fallback method also failed:\n{str(e_fallback)}\n\n"
                        "Please try using manual column selection mode."
                    )
                    self.result = None
                    return

        except Exception as e:
            messagebox.showerror("Error", f"Auto mode failed:\n{str(e)}")
            self.result = None

    def use_manual_mode(self):
        """Use manual column header selection"""
        try:
            # Save reference to parent before destroying dialog
            parent = self.dialog.master
            
            # Close this dialog
            self.dialog.destroy()

            # Open manual selection dialog
            from manual_column_selector import ManualColumnSelector

            selector = ManualColumnSelector(parent, self.file_path)
            parent.wait_window(selector.dialog)

            if selector.result:
                # Manual selection completed - cleaned file is ready
                mode, cleaned_file = selector.result
                self.result = (mode, cleaned_file)
            else:
                self.result = None

        except Exception as e:
            messagebox.showerror("Error", f"Manual mode failed:\n{str(e)}")
            self.result = None


def main():
    """Main entry point"""
    root = tk.Tk()
    app = TallyParserGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
