"""
Excel Generator - Create output Excel file from parsed tally data
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Generate formatted Excel output from tally data"""

    # Light blue fill for lengths < 30
    LIGHT_BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Light red fill for depth differences > 40
    LIGHT_RED_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    # Header style
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)

    # Border style
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, output_path):
        """
        Initialize generator

        Args:
            output_path: Path for output Excel file
        """
        self.output_path = output_path
        self.workbook = None
        self.worksheet = None

    def create(self, data, sheet_name="Tally Output"):
        """
        Create Excel file from parsed data

        Args:
            data: List of dictionaries with keys: item_num, length, depth, comment
            sheet_name: Name for the worksheet
        """
        print(f"Creating Excel file: {self.output_path}")
        print(f"Total rows: {len(data)}")

        # Create workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name

        # Write headers
        self._write_headers()

        # Write data
        self._write_data(data)

        # Format worksheet
        self._format_worksheet()

        # Save
        self.workbook.save(self.output_path)
        print(f"Done! File saved: {self.output_path}")

    def _write_headers(self):
        """Write column headers"""
        headers = ["Item #", "Length", "Depth", "Comment"]

        for col_idx, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER

    def _write_data(self, data):
        """Write data rows"""
        # Store depth cells to check differences later
        depth_cells = []
        
        for row_idx, row_data in enumerate(data, 2):  # Start from row 2 (after header)
            # Item #
            item_num_cell = self.worksheet.cell(row=row_idx, column=1, value=row_data['item_num'])
            item_num_cell.border = self.THIN_BORDER
            item_num_cell.alignment = Alignment(horizontal='center')

            # Length
            length_val = row_data['length']
            # Round to 2 decimal places if numeric
            if length_val != '' and length_val is not None:
                try:
                    length_val = round(float(length_val), 2)
                except (ValueError, TypeError):
                    pass  # Keep original value if not numeric
            length_cell = self.worksheet.cell(row=row_idx, column=2, value=length_val)
            length_cell.border = self.THIN_BORDER
            length_cell.alignment = Alignment(horizontal='right')
            length_cell.number_format = '0.00'  # Format as decimal with 2 places

            # Apply light blue fill if length < 30
            if length_val != '' and length_val is not None:
                try:
                    if float(length_val) < 30:
                        length_cell.fill = self.LIGHT_BLUE_FILL
                except (ValueError, TypeError):
                    pass  # Skip if can't convert to float

            # Depth
            depth_val = row_data['depth']
            # Round to 2 decimal places if numeric
            if depth_val != '' and depth_val is not None:
                try:
                    depth_val = round(float(depth_val), 2)
                except (ValueError, TypeError):
                    pass  # Keep original value if not numeric
            depth_cell = self.worksheet.cell(row=row_idx, column=3, value=depth_val)
            depth_cell.border = self.THIN_BORDER
            depth_cell.alignment = Alignment(horizontal='right')
            depth_cell.number_format = '0.00'  # Format as decimal with 2 places
            
            # Store depth cell and value for difference checking
            depth_cells.append((depth_cell, depth_val))

            # Comment
            comment_cell = self.worksheet.cell(row=row_idx, column=4, value=row_data['comment'])
            comment_cell.border = self.THIN_BORDER
            comment_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Check depth differences and apply light red fill
        self._highlight_large_depth_differences(depth_cells)
    
    def _highlight_large_depth_differences(self, depth_cells):
        """Highlight depth cells where difference with adjacent row is > 40"""
        for i in range(1, len(depth_cells)):
            prev_cell, prev_depth = depth_cells[i - 1]
            curr_cell, curr_depth = depth_cells[i]
            
            # Check if both depths are numeric
            try:
                prev_depth_float = float(prev_depth) if prev_depth != '' and prev_depth is not None else None
                curr_depth_float = float(curr_depth) if curr_depth != '' and curr_depth is not None else None
                
                if prev_depth_float is not None and curr_depth_float is not None:
                    # Calculate absolute difference
                    diff = abs(curr_depth_float - prev_depth_float)
                    
                    # If difference > 40, highlight both cells
                    if diff > 40:
                        prev_cell.fill = self.LIGHT_RED_FILL
                        curr_cell.fill = self.LIGHT_RED_FILL
            except (ValueError, TypeError):
                # Skip if depths are not numeric
                pass

    def _format_worksheet(self):
        """Format worksheet columns and settings"""
        # Set column widths
        self.worksheet.column_dimensions['A'].width = 12  # Item #
        self.worksheet.column_dimensions['B'].width = 12  # Length
        self.worksheet.column_dimensions['C'].width = 15  # Depth
        self.worksheet.column_dimensions['D'].width = 60  # Comment

        # Freeze header row
        self.worksheet.freeze_panes = 'A2'


def generate_excel(data, output_path, sheet_name="Tally Output"):
    """
    Convenience function to generate Excel file

    Args:
        data: List of dictionaries with tally data
        output_path: Path for output file
        sheet_name: Name for the worksheet
    """
    generator = ExcelGenerator(output_path)
    generator.create(data, sheet_name)


if __name__ == "__main__":
    import sys

    # Test data
    test_data = [
        {'item_num': '269', 'length': 36.33, 'depth': 21722.06, 'comment': 'Test comment 1'},
        {'item_num': '268', 'length': 25.5, 'depth': 21686.12, 'comment': 'Short length - should be blue'},
        {'item_num': '267', 'length': 36.71, 'depth': 21649.41, 'comment': ''},
        {'item_num': '', 'length': 18.02, 'depth': 21819.37, 'comment': 'FC Pup Joint - DV Float Collar'},
    ]

    if len(sys.argv) >= 2:
        output_path = sys.argv[1]
    else:
        output_path = "test_output.xlsx"

    generate_excel(test_data, output_path)
    print(f"\nTest file created: {output_path}")
