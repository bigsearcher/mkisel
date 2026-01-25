"""
Tally Parser - Extract data from well completion tally sheets
"""
from openpyxl import load_workbook
import re


class TallyParser:
    """Parser for well completion tally sheets"""

    # Keywords that indicate a special TYPE that should be included in comments
    SPECIAL_TYPE_KEYWORDS = [
        'DV', 'Pup', 'PUP', 'LFC', 'Float', 'Landing', 'Shoe',
        'Collar', 'Locator', 'LOCATOR', 'Packer', 'PACKER',
        'Coupling', 'COUPLING', 'Element', 'ELEMENT'
    ]

    def __init__(self, file_path):
        """
        Initialize parser with Excel file path

        Args:
            file_path: Path to cleaned Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.header_row = None
        self.column_indices = {}

    def load(self):
        """Load the Excel file and find the tally sheet"""
        print(f"Loading {self.file_path}...")
        self.workbook = load_workbook(self.file_path, data_only=True)

        # Find sheet with "Tally" but not "Deck", or use first sheet
        tally_sheet = None
        for sheet_name in self.workbook.sheetnames:
            has_tally = 'Tally' in sheet_name or 'tally' in sheet_name or 'TALLY' in sheet_name
            has_deck = 'Deck' in sheet_name or 'deck' in sheet_name or 'DECK' in sheet_name

            if has_tally and not has_deck:
                tally_sheet = sheet_name
                break

        # If no tally sheet found, try using first sheet
        if not tally_sheet:
            if len(self.workbook.sheetnames) > 0:
                tally_sheet = self.workbook.sheetnames[0]
                print(f"Warning: No 'Tally' sheet found, using first sheet: {tally_sheet}")
            else:
                raise ValueError("No sheets found in workbook")

        print(f"Using sheet: {tally_sheet}")
        self.worksheet = self.workbook[tally_sheet]

        # Find header row
        self._find_headers()

    def _find_headers(self):
        """Find the header row and column indices"""
        # Look for row containing key headers
        # Note: We'll check for LENGTH and DEPTH, JOINT/RUN is checked separately
        required_keywords = ['LENGTH', 'DEPTH']
        # Increased search range to 50 rows for more flexibility
        max_search_rows = 50

        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, max_row=max_search_rows), 1):
            row_values = [str(cell.value).upper() if cell.value else '' for cell in row]

            # Check if this row has AT LEAST ("JOINT" or "RUN") and "DEPTH" (partial match)
            has_joint = any('JOINT' in val or 'RUN' in val for val in row_values)
            has_depth = any('DEPTH' in val for val in row_values)

            if not (has_joint and has_depth):
                continue

            # This might be a header row - check next row too for combined headers
            next_row_idx = row_idx + 1
            if next_row_idx > max_search_rows:
                continue

            next_row = list(self.worksheet[next_row_idx])
            next_row_values = [str(cell.value).upper() if cell.value else '' for cell in next_row]

            # Headers might be split across two rows - combine them
            combined_headers = []
            for i, cell in enumerate(row):
                header1 = str(row_values[i]) if i < len(row_values) else ''
                header2 = str(next_row_values[i]) if i < len(next_row_values) else ''
                combined = f"{header1} {header2}".strip()
                combined_headers.append(combined)

            # Now check if combined headers have all required fields
            # Check for (JOINT or RUN) and LENGTH and DEPTH
            has_joint_or_run = any('JOINT' in val or 'RUN' in val for val in combined_headers)
            has_required = has_joint_or_run and all(any(req in val for val in combined_headers) for req in required_keywords)

            if not has_required:
                continue

            # Map column indices
            self._map_columns(combined_headers, row_idx, next_row_idx)

            if self.column_indices:
                print(f"Found headers at rows {row_idx}-{next_row_idx}")
                print(f"Column mapping: {self.column_indices}")
                return

        # Provide more detailed error message
        raise ValueError(
            "Could not find header row with required columns.\n"
            "Required columns: 'JOINT' or 'RUN', 'LENGTH', and 'DEPTH'.\n"
            "Please use manual column selection mode if your file has a different structure."
        )

    def _map_columns(self, headers, row1, row2):
        """Map column names to indices"""
        for idx, header in enumerate(headers):
            header_upper = header.upper()

            # Joint Number - check for exact matches first
            if 'JOINT' in header_upper and 'NUMBER' in header_upper and 'LENGTH' not in header_upper:
                self.column_indices['joint_number'] = idx
                self.header_row = row2  # Data starts after second header row
            elif 'JOINT' in header_upper and 'NUM' in header_upper and 'LENGTH' not in header_upper:
                self.column_indices['joint_number'] = idx
                self.header_row = row2
            # RUN NUM as joint_number (for Upper tally format)
            elif 'RUN' in header_upper and 'NUM' in header_upper and 'joint_number' not in self.column_indices:
                self.column_indices['joint_number'] = idx
                self.header_row = row2
            # Deck Number as fallback for joint_number
            elif 'DECK' in header_upper and 'NUM' in header_upper and 'joint_number' not in self.column_indices:
                self.column_indices['joint_number'] = idx
                self.header_row = row2
            # Effective Length
            elif 'EFFECTIVE' in header_upper and 'LENGTH' in header_upper:
                self.column_indices['effective_length'] = idx
                self.header_row = row2
            # Depth - must have DEPTH OF and TOP OF JT (in that order)
            elif 'DEPTH OF' in header_upper and 'TOP OF JT' in header_upper:
                # Make sure it's not a comment about depth
                if 'CALC' not in header_upper and 'SEEN' not in header_upper:
                    # Prefer this column even if we already found one
                    self.column_indices['depth'] = idx
                    self.header_row = row2
            # Fallback: just DEPTH and TOP
            elif 'depth' not in self.column_indices and 'DEPTH' in header_upper and 'TOP' in header_upper:
                if 'CALC' not in header_upper and 'SEEN' not in header_upper:
                    self.column_indices['depth'] = idx
                    self.header_row = row2
            # IN/OUT
            elif ('IN/OUT' in header_upper or 'IN OUT' in header_upper) and 'JOINT' in header_upper:
                self.column_indices['in_out'] = idx
                self.header_row = row2
            # Type
            elif 'TYPE' in header_upper and 'CENTRALIZER' not in header_upper and 'COMMENT' not in header_upper:
                self.column_indices['type'] = idx
                self.header_row = row2
            # Comments
            elif 'COMMENT' in header_upper:
                self.column_indices['comments'] = idx
                self.header_row = row2

    def _is_special_type(self, type_value):
        """Check if TYPE contains special keywords"""
        if not type_value:
            return False

        type_str = str(type_value)
        return any(keyword in type_str for keyword in self.SPECIAL_TYPE_KEYWORDS)

    def _build_comment(self, type_value, comments_value):
        """
        Build comment field according to Variant B logic:
        - If TYPE contains special keywords -> include TYPE
        - If COMMENTS has text -> include COMMENTS
        - Join with " - " separator
        """
        parts = []

        # Check if TYPE is special
        if self._is_special_type(type_value):
            type_str = str(type_value).strip()
            if type_str:
                parts.append(type_str)

        # Add COMMENTS if present
        if comments_value:
            comment_str = str(comments_value).strip()
            if comment_str and comment_str != 'None':
                parts.append(comment_str)

        return ' - '.join(parts) if parts else ''

    def parse(self):
        """
        Parse the tally sheet and extract data

        Returns:
            List of dictionaries with keys: item_num, length, depth, comment
        """
        if not self.worksheet or not self.header_row:
            raise ValueError("Must call load() first")

        data = []
        data_start_row = self.header_row + 1

        print(f"Parsing data starting from row {data_start_row}...")

        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=data_start_row), data_start_row):
            values = [cell.value for cell in row]

            # Check IN/OUT if column exists
            in_out_idx = self.column_indices.get('in_out')
            if in_out_idx is not None:
                # Only include rows where IN/OUT = "IN" if column exists
                if in_out_idx >= len(values):
                    continue

                in_out = values[in_out_idx]
                if str(in_out).strip().upper() != 'IN':
                    continue
            # If no IN/OUT column, include all rows (but check depth below)

            # Extract values
            joint_num_idx = self.column_indices.get('joint_number', 3)
            length_idx = self.column_indices.get('effective_length', 5)
            depth_idx = self.column_indices.get('depth', 8)
            type_idx = self.column_indices.get('type', 2)
            comments_idx = self.column_indices.get('comments', 10)

            joint_number = values[joint_num_idx] if joint_num_idx < len(values) else None
            length = values[length_idx] if length_idx < len(values) else None
            depth = values[depth_idx] if depth_idx < len(values) else None
            type_val = values[type_idx] if type_idx < len(values) else None
            comments_val = values[comments_idx] if comments_idx < len(values) else None

            # Skip if no depth
            if depth is None:
                continue

            # Skip if depth is not a number (e.g., "Top Collar", headers, etc.)
            try:
                float(depth)
            except (ValueError, TypeError):
                continue

            # Build comment
            comment = self._build_comment(type_val, comments_val)

            # Add to data
            data.append({
                'item_num': joint_number if joint_number is not None else '',
                'length': length if length is not None else '',
                'depth': depth if depth is not None else '',
                'comment': comment
            })

        print(f"Extracted {len(data)} rows")

        # Sort by depth (ascending)
        def safe_float(x):
            try:
                return float(x['depth']) if x['depth'] != '' else 0
            except (ValueError, TypeError):
                return 0

        data.sort(key=safe_float)
        print("Data sorted by depth (ascending)")

        return data


def parse_tally_file(file_path):
    """
    Convenience function to parse a tally file

    Args:
        file_path: Path to cleaned Excel file

    Returns:
        List of dictionaries with tally data
    """
    parser = TallyParser(file_path)
    parser.load()
    return parser.parse()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python tally_parser.py <cleaned_excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        data = parse_tally_file(file_path)

        print("\n" + "="*80)
        print("PARSED DATA (first 10 rows):")
        print("="*80)

        for i, row in enumerate(data[:10], 1):
            print(f"\nRow {i}:")
            print(f"  Item #:  {row['item_num']}")
            print(f"  Length:  {row['length']}")
            print(f"  Depth:   {row['depth']}")
            print(f"  Comment: {row['comment']}")

        if len(data) > 10:
            print(f"\n... and {len(data) - 10} more rows")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
