"""
Manual Column Selector - Step-by-step dialog for manual column header selection
"""
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from pathlib import Path
import os
import re
from tkinter.font import Font


class ManualColumnSelector:
    """Step-by-step dialog for manually selecting column headers from Excel file"""

    def __init__(self, parent, file_path):
        self.file_path = file_path
        self.result = None
        self.workbook = None
        self.worksheet = None

        # Selected values
        self.selected_header_row = None
        self.selected_depth_col = None
        self.selected_length_col = None
        self.selected_item_col = None
        self.selected_comments_col = None
        
        # Track temporary files for cleanup
        self._temp_file_to_cleanup = None

        # Current step
        self.current_step = 0
        self.steps = [
            "header_row",
            "depth",
            "length",
            "item_number",
            "comments"
        ]

        # Create dialog first (needed for messagebox)
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Manual Column Selection")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        # Bind window close event to cleanup
        self.dialog.protocol("WM_DELETE_WINDOW", self._cleanup_and_close)

        # Hide dialog initially until data is loaded
        self.dialog.withdraw()
        
        # Load Excel data
        try:
            self._load_excel_data()
            
            if not self.worksheet:
                self._cleanup_and_close()
                return

            # Auto-detect header row
            self._auto_detect_header_row()

            # Setup UI
            self._setup_ui()
            
            # Show dialog and maximize to available work area (accounts for taskbar)
            self.dialog.deiconify()
            self.dialog.update_idletasks()
            
            # Get available work area (screen minus taskbar)
            # For Windows, we need to get the work area
            try:
                # Try to get work area using Windows API
                import ctypes
                from ctypes import wintypes
                
                user32 = ctypes.windll.user32
                # Get work area (screen minus taskbar)
                class RECT(ctypes.Structure):
                    _fields_ = [("left", ctypes.c_long),
                              ("top", ctypes.c_long),
                              ("right", ctypes.c_long),
                              ("bottom", ctypes.c_long)]
                
                rect = RECT()
                user32.SystemParametersInfoW(48, 0, ctypes.byref(rect), 0)  # SPI_GETWORKAREA = 48
                
                work_width = rect.right - rect.left
                work_height = rect.bottom - rect.top
                work_x = rect.left
                work_y = rect.top
                
                # Set window size to work area
                self.dialog.geometry(f"{work_width}x{work_height}+{work_x}+{work_y}")
            except Exception:
                # Fallback: use screen size minus estimated taskbar height
                screen_width = self.dialog.winfo_screenwidth()
                screen_height = self.dialog.winfo_screenheight()
                # Estimate taskbar height (usually 40-50 pixels)
                work_height = screen_height - 50
                self.dialog.geometry(f"{screen_width}x{work_height}+0+0")
            
            self.dialog.update_idletasks()
        except Exception as e:
            # If loading fails, show error and close dialog
            if hasattr(self, 'dialog'):
                self._cleanup_and_close()
            messagebox.showerror("Error", f"Failed to initialize manual column selector:\n{str(e)}")
            return

    def _load_excel_data(self):
        """Load first 30 rows from Excel file"""
        try:
            file_path = Path(self.file_path)
            
            # If file is .xls, convert to .xlsx first using pandas + xlrd
            if file_path.suffix.lower() == '.xls':
                import pandas as pd
                try:
                    # Read .xls file using pandas with xlrd engine
                    all_sheets = pd.read_excel(self.file_path, sheet_name=None, engine='xlrd')
                    
                    # Create .xlsx file path
                    xlsx_file = file_path.parent / f"{file_path.stem}_converted.xlsx"
                    
                    # Delete existing converted file if it exists
                    if xlsx_file.exists():
                        try:
                            xlsx_file.unlink()
                        except Exception:
                            pass
                    
                    # Write to .xlsx using openpyxl engine
                    with pd.ExcelWriter(str(xlsx_file), engine='openpyxl') as writer:
                        for sheet_name, df in all_sheets.items():
                            # Truncate sheet name if too long (Excel limit is 31 characters)
                            sheet_name_truncated = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                            df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)
                    
                    # Use converted file
                    self.file_path = str(xlsx_file)
                    file_path = xlsx_file
                    
                except ImportError as e:
                    messagebox.showerror(
                        "Missing Package",
                        f"xlrd package is required to convert .xls files.\n\n"
                        f"Please install it:\npip install xlrd\n\n"
                        f"Error: {str(e)}"
                    )
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to convert .xls to .xlsx:\n{str(e)}")
                    return
            
            # Try to load workbook
            try:
                self.workbook = load_workbook(self.file_path, data_only=True)
            except Exception as load_error:
                # If loading fails due to invalid XML, try to fix it using Excel COM
                error_msg = str(load_error).lower()
                if 'xml' in error_msg or 'unable to read workbook' in error_msg or 'could not assign names' in error_msg:
                    # Check if dialog exists before showing messagebox
                    if not hasattr(self, 'dialog') or not self.dialog.winfo_exists():
                        # Dialog doesn't exist, can't show messagebox - just raise error
                        raise load_error
                    
                    # Ask user if they want to try fixing the file
                    try:
                        response = messagebox.askyesno(
                            "Error Loading File",
                            f"Failed to load Excel file:\n{str(load_error)}\n\n"
                            "The file may contain invalid XML or corrupted data.\n\n"
                            "Would you like to try fixing it in Excel?\n"
                            "(This will open the file in Excel, calculate formulas, and save it)",
                            parent=self.dialog
                        )
                    except Exception as msg_error:
                        # If messagebox fails, just raise the original error
                        raise load_error
                    
                    if response:
                        try:
                            from utils.calculate_formulas import calculate_formulas
                            
                            # Create temporary fixed file
                            temp_fixed = file_path.parent / f"{file_path.stem}_fixed_temp.xlsx"
                            
                            # Fix file using Excel COM
                            calculate_formulas(str(file_path), str(temp_fixed))
                            
                            # Try loading the fixed file
                            self.file_path = str(temp_fixed)
                            self.workbook = load_workbook(self.file_path, data_only=True)
                            
                            # Track temporary file for cleanup
                            self._temp_file_to_cleanup = temp_fixed
                        except Exception as fix_error:
                            if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                                messagebox.showerror(
                                    "Error",
                                    f"Failed to fix file:\n{str(fix_error)}\n\n"
                                    "Please try opening the file in Excel manually, "
                                    "saving it, and then try again.",
                                    parent=self.dialog
                                )
                            raise load_error
                    else:
                        raise load_error
                else:
                    raise load_error

            # Find first sheet (or Tally sheet)
            tally_sheet = None
            for sheet_name in self.workbook.sheetnames:
                has_tally = 'Tally' in sheet_name or 'tally' in sheet_name or 'TALLY' in sheet_name
                has_deck = 'Deck' in sheet_name or 'deck' in sheet_name or 'DECK' in sheet_name

                if has_tally and not has_deck:
                    tally_sheet = sheet_name
                    break

            if not tally_sheet and len(self.workbook.sheetnames) > 0:
                tally_sheet = self.workbook.sheetnames[0]

            self.worksheet = self.workbook[tally_sheet]

        except Exception as e:
            # Clean up temporary file if it exists
            if hasattr(self, '_temp_file_to_cleanup') and self._temp_file_to_cleanup:
                try:
                    if self._temp_file_to_cleanup.exists():
                        self._temp_file_to_cleanup.unlink()
                except Exception:
                    pass


            messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}")
            if hasattr(self, 'dialog'):
                self._cleanup_and_close()

    def _auto_detect_header_row(self):
        """Auto-detect header row by searching for keywords DEPTH, EFFECTIVE, COMMENT"""
        if not self.worksheet:
            return
        
        # Keywords to search for
        keywords = ['DEPTH', 'EFFECTIVE', 'COMMENT', 'COMMENTS']
        
        # Search in first 30 rows
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            row_values = [str(cell.value or "").upper() for cell in row]
            
            # Check if this row contains at least 2 keywords
            found_keywords = sum(1 for keyword in keywords if any(keyword in val for val in row_values))
            
            if found_keywords >= 2:
                # Also check next row (headers might be split across two rows)
                if row_idx < 30:
                    next_row = list(self.worksheet[row_idx + 1])
                    next_row_values = [str(cell.value or "").upper() for cell in next_row]
                    combined_found = sum(1 for keyword in keywords if any(keyword in val for val in row_values + next_row_values))
                    if combined_found >= 2:
                        self.selected_header_row = row_idx
                        return
                
                self.selected_header_row = row_idx
                return

    def _auto_detect_column(self, column_type):
        """Auto-detect column by searching for keywords"""
        if not self.worksheet or not self.selected_header_row:
            return None
        
        # Map column types to keywords
        keyword_map = {
            'depth': ['DEPTH', 'TOP OF JT', 'TOP OF JOINT'],
            'length': ['EFFECTIVE', 'LENGTH', 'EFF LENGTH'],
            'item_number': ['ITEM', 'JOINT', 'RUN NUM', 'JOINT NUMBER', 'JOINT NUM'],
            'comments': ['COMMENT', 'COMMENTS', 'NOTE', 'NOTES', 'DESCRIPTION']
        }
        
        keywords = keyword_map.get(column_type, [])
        if not keywords:
            return None
        
        # Check header row and next row (headers might be split)
        header_row = list(self.worksheet[self.selected_header_row])
        next_row = list(self.worksheet[self.selected_header_row + 1]) if self.selected_header_row < 30 else []
        
        best_match = None
        best_score = 0
        
        max_cols = max(len(header_row), len(next_row) if next_row else 0)
        
        for col_idx in range(max_cols):
            # Combine header values from both rows
            header_val = str(header_row[col_idx].value or "").upper() if col_idx < len(header_row) else ""
            next_val = str(next_row[col_idx].value or "").upper() if next_row and col_idx < len(next_row) else ""
            combined_val = f"{header_val} {next_val}".strip()
            
            # Score based on keyword matches
            score = sum(1 for keyword in keywords if keyword in combined_val)
            
            if score > best_score:
                best_score = score
                best_match = col_idx
        
        return best_match if best_score > 0 else None

    def _setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title frame
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 10))

        self.title_label = ttk.Label(
            title_frame,
            text="",
            font=('Arial', 14, 'bold')
        )
        self.title_label.pack()

        self.instruction_label = ttk.Label(
            title_frame,
            text="",
            font=('Arial', 10),
            foreground="gray"
        )
        self.instruction_label.pack(pady=(5, 0))

        # Table frame
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self._create_table(table_frame)

        # Button frame - centered
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)

        # Center buttons
        button_container = ttk.Frame(button_frame)
        button_container.pack(expand=True)

        # Create style for larger buttons
        style = ttk.Style()
        style.configure('Large.TButton', padding=(20, 10))
        
        self.confirm_button = ttk.Button(
            button_container,
            text="",
            command=self._confirm_step,
            width=30,
            style='Large.TButton'
        )
        self.confirm_button.pack(side=tk.LEFT, padx=10)

        self.cancel_button = ttk.Button(
            button_container,
            text="Cancel",
            command=self._cleanup_and_close,
            width=30,
            style='Large.TButton'
        )
        self.cancel_button.pack(side=tk.LEFT, padx=10)

        # Status label
        self.status_label = ttk.Label(
            main_frame,
            text="",
            font=('Arial', 9),
            foreground="blue"
        )
        self.status_label.pack(pady=(5, 0))

        # Initialize first step
        self._show_step()

    def _create_table(self, parent):
        """Create table using Canvas for better cell selection"""
        # Create frame for table with scrollbars
        table_container = ttk.Frame(parent)
        table_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbars
        vsb = ttk.Scrollbar(table_container, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        hsb = ttk.Scrollbar(table_container, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # Canvas for table
        self.table_canvas = tk.Canvas(
            table_container,
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            bg='white',
            highlightthickness=0
        )
        self.table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb.config(command=self.table_canvas.yview)
        hsb.config(command=self.table_canvas.xview)

        # Create frame inside canvas for table content
        self.table_frame = ttk.Frame(self.table_canvas)
        self.table_window = self.table_canvas.create_window((0, 0), window=self.table_frame, anchor="nw")

        # Bind canvas scrolling
        self.table_frame.bind('<Configure>', self._on_table_frame_configure)
        self.table_canvas.bind('<Configure>', self._on_canvas_configure)
        
        # Store table data
        self.table_data = []  # Will store [row][col] = value
        self.table_cells = {}  # Will store (row, col) -> widget reference
        self.cell_width = 120
        self.cell_height = 25
        self.header_height = 30
        
        # Store last clicked cell
        self.last_clicked_row = None
        self.last_clicked_col = None
        
        # Selected column for highlighting
        self.selected_column = None
        
        # Load data into table
        self._populate_table()
        
        # Bind mouse events
        self.table_canvas.bind('<Button-1>', self._on_canvas_click)
    
    def _on_table_frame_configure(self, event):
        """Update scroll region when table frame size changes"""
        self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all"))
    
    def _on_canvas_configure(self, event):
        """Update canvas window size"""
        canvas_width = event.width
        canvas_height = event.height
        self.table_canvas.itemconfig(self.table_window, width=canvas_width, height=canvas_height)

    def _populate_table(self):
        """Populate table with first 30 rows using Canvas-based table"""
        if not self.worksheet:
            return

        # Clear existing widgets
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        self.table_data = []
        self.table_cells = {}

        # Get max columns (check first 30 rows)
        max_cols = 0
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            max_cols = max(max_cols, len(row))

        # Create header row
        row_num_label = ttk.Label(
            self.table_frame,
            text="Row",
            font=('Arial', 9, 'bold'),
            background='#E0E0E0',
            relief=tk.RAISED,
            borderwidth=1,
            width=8,
            anchor='center'
        )
        row_num_label.grid(row=0, column=0, sticky='nsew', padx=1, pady=1)

        # Create column headers (A, B, C, ...)
        for col_idx in range(max_cols):
            col_letter = self._number_to_column_letter(col_idx + 1)
            header_label = ttk.Label(
                self.table_frame,
                text=col_letter,
                font=('Arial', 9, 'bold'),
                background='#E0E0E0',
                relief=tk.RAISED,
                borderwidth=1,
                width=15,
                anchor='center'
            )
            header_label.grid(row=0, column=col_idx + 1, sticky='nsew', padx=1, pady=1)

        # Add data rows
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            row_data = []
            row_cells = []

            # Row number label
            row_num_label = ttk.Label(
                self.table_frame,
                text=str(row_idx),
                font=('Arial', 9),
                background='#F0F0F0',
                relief=tk.SUNKEN,
                borderwidth=1,
                width=8,
                anchor='center'
            )
            row_num_label.grid(row=row_idx, column=0, sticky='nsew', padx=1, pady=1)
            row_cells.append(row_num_label)

            # Data cells
            for col_idx in range(max_cols):
                if col_idx < len(row):
                    cell_value = row[col_idx].value
                    if cell_value is None:
                        cell_text = ''
                    else:
                        cell_text = str(cell_value)
                        if len(cell_text) > 50:
                            cell_text = cell_text[:47] + '...'
                else:
                    cell_text = ''

                row_data.append(cell_text)

                # Create cell label
                cell_label = ttk.Label(
                    self.table_frame,
                    text=cell_text,
                    font=('Arial', 9),
                    background='white',
                    relief=tk.SUNKEN,
                    borderwidth=1,
                    width=15,
                    anchor='w',
                    padding=(3, 2)
                )
                cell_label.grid(row=row_idx, column=col_idx + 1, sticky='nsew', padx=1, pady=1)
                
                # Bind click event to cell
                cell_label.bind('<Button-1>', lambda e, r=row_idx, c=col_idx: self._on_cell_click_canvas(r, c))
                
                # Store cell reference
                self.table_cells[(row_idx, col_idx)] = cell_label
                row_cells.append(cell_label)

            self.table_data.append(row_data)

        # Configure grid weights for resizing
        self.table_frame.columnconfigure(0, weight=0, minsize=60)
        for col_idx in range(max_cols):
            self.table_frame.columnconfigure(col_idx + 1, weight=1, minsize=120)
        
        # Update scroll region
        self.table_canvas.update_idletasks()
        self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all"))
    
    def _on_cell_click_canvas(self, row, col):
        """Handle cell click in Canvas-based table"""
        step_name = self.steps[self.current_step]
        
        if step_name == "header_row":
            # For header row selection, select the entire row
            self.last_clicked_row = row
            self._clear_all_highlights()
            self._highlight_row_canvas(row)
            col_letter = self._number_to_column_letter(col + 1)
            self.status_label.config(
                text=f"Selected: Row {row}, Column {col_letter}",
                foreground="green"
            )
        else:
            # For column selection, select the entire column
            self.last_clicked_col = col
            self._clear_all_highlights()
            self._highlight_column_canvas(col)
            col_letter = self._number_to_column_letter(col + 1)
            self.status_label.config(
                text=f"Selected: Column {col_letter} (Row {row})",
                foreground="green"
            )
    
    def _on_canvas_click(self, event):
        """Handle clicks on canvas (for scrolling)"""
        pass

    def _number_to_column_letter(self, n):
        """Convert column number to letter (1=A, 2=B, etc.)"""
        result = ""
        while n > 0:
            n -= 1
            result = chr(n % 26 + 65) + result
            n //= 26
        return result
    
    def _clear_all_highlights(self):
        """Clear all highlights from Canvas-based table"""
        # Reset all cells to default background
        for (row, col), cell_label in self.table_cells.items():
            if row == 0:  # Header row
                cell_label.config(background='#E0E0E0')
            else:
                cell_label.config(background='white')
        
        # Also reset row number labels
        for widget in self.table_frame.winfo_children():
            if isinstance(widget, ttk.Label) and widget.grid_info().get('column') == 0:
                row_num = widget.grid_info().get('row')
                if row_num > 0:
                    widget.config(background='#F0F0F0')
    
    def _highlight_row_canvas(self, row_num):
        """Highlight entire row in Canvas-based table"""
        # Highlight row number label
        for widget in self.table_frame.winfo_children():
            grid_info = widget.grid_info()
            if grid_info.get('row') == row_num and grid_info.get('column') == 0:
                widget.config(background='#90EE90')
                break
        
        # Highlight all cells in the row
        for col_idx in range(len(self.table_data[row_num - 1]) if row_num <= len(self.table_data) else 0):
            if (row_num, col_idx) in self.table_cells:
                self.table_cells[(row_num, col_idx)].config(background='#90EE90')
        
        # Scroll to row
        self.table_canvas.update_idletasks()
        try:
            row_widget = self.table_cells.get((row_num, 0))
            if row_widget:
                self.table_canvas.see(row_widget)
        except:
            pass
    
    def _highlight_column_canvas(self, col_index):
        """Highlight entire column in Canvas-based table - only the selected column cells"""
        # Highlight column header
        for widget in self.table_frame.winfo_children():
            grid_info = widget.grid_info()
            if grid_info.get('row') == 0 and grid_info.get('column') == col_index + 1:
                widget.config(background='#FFD700')  # Gold for header
                break
        
        # Highlight only cells in the selected column (not entire rows)
        for row_idx in range(1, len(self.table_data) + 1):
            if (row_idx, col_index) in self.table_cells:
                self.table_cells[(row_idx, col_index)].config(background='#E6F3FF')  # Light blue for cells
        
        # Scroll to column
        self.table_canvas.update_idletasks()
        try:
            # Scroll to show first cell of the column
            if (1, col_index) in self.table_cells:
                cell_widget = self.table_cells[(1, col_index)]
                self.table_canvas.see(cell_widget)
        except:
            pass
    
    def _highlight_row(self, row_num):
        """Highlight entire row (wrapper for Canvas)"""
        self._highlight_row_canvas(row_num)
    
    def _highlight_column(self, col_index):
        """Highlight entire column (wrapper for Canvas)"""
        self._highlight_column_canvas(col_index)

    def _show_step(self):
        """Show current step in the same table"""
        step_name = self.steps[self.current_step]
        
        # Clear all highlights
        self._clear_all_highlights()
        
        if step_name == "header_row":
            # Step 1: Select header row
            self.title_label.config(text="Choose Table Header Row")
            self.instruction_label.config(
                text="Click on a cell in the header row that contains column names (Item, Depth, Length, etc.)"
            )
            self.confirm_button.config(text="Confirm Header Row")
            
            # Highlight auto-detected row if available
            if self.selected_header_row:
                self._highlight_row(self.selected_header_row)
                col_letter = self._number_to_column_letter(1)
                self.status_label.config(
                    text=f"Auto-detected: Row {self.selected_header_row}. Click to change or click 'Confirm Header Row' to confirm.",
                    foreground="green"
                )
            else:
                self.status_label.config(
                    text="Click on a cell in the header row",
                    foreground="blue"
                )
                
        elif step_name == "depth":
            # Step 2: Select depth column
            self.title_label.config(text="Choose Depth Column")
            self.instruction_label.config(text="Click on a cell in the depth column")
            self.confirm_button.config(text="Choose Depth")
            
            # Highlight auto-detected column if available
            auto_col = self._auto_detect_column("depth")
            if auto_col is not None and self.selected_depth_col is None:
                self.selected_depth_col = auto_col
                self._highlight_column(auto_col)
                col_letter = self._number_to_column_letter(auto_col + 1)
                self.status_label.config(
                    text=f"Auto-detected: Column {col_letter}. Click to change or click 'Choose Depth' to confirm.",
                    foreground="green"
                )
            elif self.selected_depth_col is not None:
                self._highlight_column(self.selected_depth_col)
                col_letter = self._number_to_column_letter(self.selected_depth_col + 1)
                self.status_label.config(
                    text=f"Selected: Column {col_letter}. Click to change or click 'Choose Depth' to confirm.",
                    foreground="green"
                )
            else:
                self.status_label.config(
                    text="Click on a cell in the depth column",
                    foreground="blue"
                )
                
        elif step_name == "length":
            # Step 3: Select length column
            self.title_label.config(text="Choose Effective Length Column")
            self.instruction_label.config(text="Click on a cell in the effective length column")
            self.confirm_button.config(text="Choose Length")
            
            auto_col = self._auto_detect_column("length")
            if auto_col is not None and self.selected_length_col is None:
                self.selected_length_col = auto_col
                self._highlight_column(auto_col)
                col_letter = self._number_to_column_letter(auto_col + 1)
                self.status_label.config(
                    text=f"Auto-detected: Column {col_letter}. Click to change or click 'Choose Length' to confirm.",
                    foreground="green"
                )
            elif self.selected_length_col is not None:
                self._highlight_column(self.selected_length_col)
                col_letter = self._number_to_column_letter(self.selected_length_col + 1)
                self.status_label.config(
                    text=f"Selected: Column {col_letter}. Click to change or click 'Choose Length' to confirm.",
                    foreground="green"
                )
            else:
                self.status_label.config(
                    text="Click on a cell in the effective length column",
                    foreground="blue"
                )
                
        elif step_name == "item_number":
            # Step 4: Select item number column
            self.title_label.config(text="Choose Item Number Column")
            self.instruction_label.config(text="Click on a cell in the item number column")
            self.confirm_button.config(text="Choose Item Number")
            
            auto_col = self._auto_detect_column("item_number")
            if auto_col is not None and self.selected_item_col is None:
                self.selected_item_col = auto_col
                self._highlight_column(auto_col)
                col_letter = self._number_to_column_letter(auto_col + 1)
                self.status_label.config(
                    text=f"Auto-detected: Column {col_letter}. Click to change or click 'Choose Item Number' to confirm.",
                    foreground="green"
                )
            elif self.selected_item_col is not None:
                self._highlight_column(self.selected_item_col)
                col_letter = self._number_to_column_letter(self.selected_item_col + 1)
                self.status_label.config(
                    text=f"Selected: Column {col_letter}. Click to change or click 'Choose Item Number' to confirm.",
                    foreground="green"
                )
            else:
                self.status_label.config(
                    text="Click on a cell in the item number column (optional)",
                    foreground="blue"
                )
                
        elif step_name == "comments":
            # Step 5: Select comments column
            self.title_label.config(text="Choose Comments Column")
            self.instruction_label.config(text="Click on a cell in the comments column")
            self.confirm_button.config(text="Choose Comments")
            
            auto_col = self._auto_detect_column("comments")
            if auto_col is not None and self.selected_comments_col is None:
                self.selected_comments_col = auto_col
                self._highlight_column(auto_col)
                col_letter = self._number_to_column_letter(auto_col + 1)
                self.status_label.config(
                    text=f"Auto-detected: Column {col_letter}. Click to change or click 'Choose Comments' to confirm.",
                    foreground="green"
                )
            elif self.selected_comments_col is not None:
                self._highlight_column(self.selected_comments_col)
                col_letter = self._number_to_column_letter(self.selected_comments_col + 1)
                self.status_label.config(
                    text=f"Selected: Column {col_letter}. Click to change or click 'Choose Comments' to confirm.",
                    foreground="green"
                )
            else:
                self.status_label.config(
                    text="Click on a cell in the comments column (optional)",
                    foreground="blue"
                )

    def _scroll_to_row(self, row_num):
        """Scroll table to show specific row"""
        try:
            for item in self.table.get_children():
                if self.table.item(item, 'text') == str(row_num):
                    self.table.see(item)
                    self.table.selection_set(item)
                    break
        except Exception:
            pass

    def _scroll_to_column(self, col_idx):
        """Scroll table to show specific column"""
        try:
            # Select first visible row in that column
            children = self.table.get_children()
            if children:
                self.table.selection_set(children[0])
                self.table.see(children[0])
        except Exception:
            pass

    def _show_column_selection_dialog_OLD(self, column_type, dialog_title, column_name):
        """Show separate dialog for column selection"""
        # Hide main dialog
        self.dialog.withdraw()
        
        # Create modal dialog
        col_dialog = tk.Toplevel(self.dialog)
        col_dialog.title(dialog_title)
        col_dialog.transient(self.dialog)
        col_dialog.grab_set()
        
        # Main frame
        main_frame = ttk.Frame(col_dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text=dialog_title,
            font=('Arial', 14, 'bold')
        )
        title_label.pack(pady=(0, 10))
        
        # Instruction
        instruction_label = ttk.Label(
            main_frame,
            text=f"Click on a cell in the {column_name} column",
            font=('Arial', 10),
            foreground="gray"
        )
        instruction_label.pack(pady=(0, 10))
        
        # Table frame
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Create table (reuse existing table creation logic)
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        
        table = ttk.Treeview(
            table_frame,
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            selectmode='none'  # Disable row selection - we only want cell clicks
        )
        table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        vsb.config(command=table.yview)
        hsb.config(command=table.xview)
        
        # Populate table
        max_cols = 0
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            max_cols = max(max_cols, len(row))
        
        columns = [f"Col{i}" for i in range(max_cols)]
        table['columns'] = columns
        table['show'] = 'tree headings'
        
        table.heading('#0', text='Row')
        table.column('#0', width=60, stretch=False)
        
        for i, col in enumerate(columns):
            col_letter = self._number_to_column_letter(i + 1)
            table.heading(col, text=col_letter)
            table.column(col, width=120, stretch=True)
        
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            values = []
            
            for cell in row:
                value = cell.value
                if value is None:
                    values.append('')
                else:
                    str_value = str(value)
                    if len(str_value) > 50:
                        str_value = str_value[:47] + '...'
                    values.append(str_value)
            
            while len(values) < max_cols:
                values.append('')
            
            table.insert('', 'end', text=str(row_idx), values=values)
        
        # Status label
        status_label = ttk.Label(
            main_frame,
            text="",
            font=('Arial', 9),
            foreground="blue"
        )
        status_label.pack(pady=(5, 0))
        
        # Selected column
        selected_col = [None]  # Use list to allow modification in nested function
        
        # Auto-detect column
        auto_col = self._auto_detect_column(column_type)
        if auto_col is not None:
            selected_col[0] = auto_col
            col_letter = self._number_to_column_letter(auto_col + 1)
            status_label.config(
                text=f"Auto-detected: Column {col_letter}. Click to change or click 'Choose {column_name.title()}' to confirm.",
                foreground="green"
            )
            # Highlight entire auto-detected column - clear any previous highlights first
            all_items = table.get_children()
            for item_row in all_items:
                tags = list(table.item(item_row, 'tags'))
                if 'highlighted_col' in tags:
                    tags.remove('highlighted_col')
                table.item(item_row, tags=tuple(tags))
            # Highlight all cells in auto-detected column
            for item_row in all_items:
                tags = list(table.item(item_row, 'tags'))
                tags.append('highlighted_col')
                table.item(item_row, tags=tuple(tags))
            # Scroll to column
            if all_items:
                table.see(all_items[0])
        else:
            status_label.config(text="Click on a cell in the column")
        
        # Configure tag for highlighting column
        table.tag_configure("highlighted_col", background="#E6F3FF")  # Light blue for highlighted column
        
        # Handle cell click
        def on_cell_click(event):
            item = table.identify_row(event.y)
            if not item:
                return
            
            column = table.identify_column(event.x)
            if not column or column == '#0':
                return
            
            row_text = table.item(item, 'text')
            try:
                row_num = int(row_text)
            except:
                return
            
            col_index = int(column.replace('#', '')) - 1
            selected_col[0] = col_index
            col_letter = self._number_to_column_letter(col_index + 1)
            status_label.config(
                text=f"Selected: Column {col_letter} (Row {row_num})",
                foreground="green"
            )
            
            # Highlight entire column - clear previous highlights from all rows
            all_items = table.get_children()
            for item_row in all_items:
                tags = list(table.item(item_row, 'tags'))
                if 'highlighted_col' in tags:
                    tags.remove('highlighted_col')
                table.item(item_row, tags=tuple(tags))
            
            # Highlight all cells in selected column (entire column)
            for item_row in all_items:
                tags = list(table.item(item_row, 'tags'))
                tags.append('highlighted_col')
                table.item(item_row, tags=tuple(tags))
        
        table.bind('<Button-1>', on_cell_click)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        def choose_column():
            if selected_col[0] is None:
                messagebox.showwarning("Warning", f"Please click on a cell in the {column_name} column first")
                return
            
            # Set the selected column
            if column_type == "depth":
                self.selected_depth_col = selected_col[0]
            elif column_type == "length":
                self.selected_length_col = selected_col[0]
            elif column_type == "item_number":
                self.selected_item_col = selected_col[0]
            elif column_type == "comments":
                self.selected_comments_col = selected_col[0]
            
            col_dialog.destroy()
            # Show main dialog again
            self.dialog.deiconify()
            self.current_step += 1
            self.last_clicked_col = None
            self._show_step()
        
        choose_button = ttk.Button(
            button_frame,
            text=f"Choose {column_name.title()}",
            command=choose_column,
            width=20
        )
        choose_button.pack(side=tk.LEFT, padx=5)
        
        cancel_button = ttk.Button(
            button_frame,
            text="Cancel",
            command=col_dialog.destroy,
            width=20
        )
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        # Make window fullscreen after all widgets are created
        col_dialog.update_idletasks()
        col_dialog.state('zoomed')  # Windows - maximize to fullscreen
        # For Linux/Mac, use: col_dialog.attributes('-zoomed', True)

    def _show_header_row_selection_dialog_OLD(self):
        """Show separate dialog for header row selection"""
        # Hide main dialog
        self.dialog.withdraw()
        
        # Create modal dialog
        header_dialog = tk.Toplevel(self.dialog)
        header_dialog.title("Choose Table Header Row")
        header_dialog.transient(self.dialog)
        header_dialog.grab_set()
        
        # Main frame
        main_frame = ttk.Frame(header_dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text="Choose Table Header Row",
            font=('Arial', 14, 'bold')
        )
        title_label.pack(pady=(0, 10))
        
        # Instruction
        instruction_label = ttk.Label(
            main_frame,
            text="Click on a cell in the header row that contains column names (Item, Depth, Length, etc.)",
            font=('Arial', 10),
            foreground="gray"
        )
        instruction_label.pack(pady=(0, 10))
        
        # Table frame
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Create table
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        
        table = ttk.Treeview(
            table_frame,
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            selectmode='none'
        )
        table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        vsb.config(command=table.yview)
        hsb.config(command=table.xview)
        
        # Populate table
        max_cols = 0
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            max_cols = max(max_cols, len(row))
        
        columns = [f"Col{i}" for i in range(max_cols)]
        table['columns'] = columns
        table['show'] = 'tree headings'
        
        table.heading('#0', text='Row')
        table.column('#0', width=60, stretch=False)
        
        for i, col in enumerate(columns):
            col_letter = self._number_to_column_letter(i + 1)
            table.heading(col, text=col_letter)
            table.column(col, width=120, stretch=True)
        
        for row_idx in range(1, min(31, self.worksheet.max_row + 1)):
            row = list(self.worksheet[row_idx])
            values = []
            
            for cell in row:
                value = cell.value
                if value is None:
                    values.append('')
                else:
                    str_value = str(value)
                    if len(str_value) > 50:
                        str_value = str_value[:47] + '...'
                    values.append(str_value)
            
            while len(values) < max_cols:
                values.append('')
            
            table.insert('', 'end', text=str(row_idx), values=values)
        
        # Configure tag for highlighting row
        table.tag_configure("highlighted_row", background="#E6F3FF")
        
        # Status label
        status_label = ttk.Label(
            main_frame,
            text="",
            font=('Arial', 9),
            foreground="blue"
        )
        status_label.pack(pady=(5, 0))
        
        # Selected row
        selected_row = [self.selected_header_row]  # Start with auto-detected or None
        
        # Highlight auto-detected row if available
        if self.selected_header_row:
            for item in table.get_children():
                if table.item(item, 'text') == str(self.selected_header_row):
                    tags = list(table.item(item, 'tags'))
                    tags.append('highlighted_row')
                    table.item(item, tags=tuple(tags))
                    table.see(item)
                    break
            status_label.config(
                text=f"Auto-detected: Row {self.selected_header_row}. Click to change or click 'Confirm Header Row' to confirm.",
                foreground="green"
            )
        else:
            status_label.config(text="Click on a cell in the header row")
        
        # Handle cell click
        def on_cell_click(event):
            item = table.identify_row(event.y)
            if not item:
                return
            
            row_text = table.item(item, 'text')
            try:
                row_num = int(row_text)
            except:
                return
            
            selected_row[0] = row_num
            
            # Clear previous highlights
            for item_row in table.get_children():
                tags = list(table.item(item_row, 'tags'))
                if 'highlighted_row' in tags:
                    tags.remove('highlighted_row')
                table.item(item_row, tags=tuple(tags))
            
            # Highlight selected row
            for item_row in table.get_children():
                if table.item(item_row, 'text') == str(row_num):
                    tags = list(table.item(item_row, 'tags'))
                    tags.append('highlighted_row')
                    table.item(item_row, tags=tuple(tags))
                    table.see(item_row)
                    break
            
            status_label.config(
                text=f"Selected: Row {row_num}",
                foreground="green"
            )
        
        table.bind('<Button-1>', on_cell_click)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        def confirm_header_row():
            if selected_row[0] is None:
                messagebox.showwarning("Warning", "Please click on a cell in the header row first")
                return
            
            self.selected_header_row = selected_row[0]
            header_dialog.destroy()
            # Show main dialog again
            self.dialog.deiconify()
            self.current_step += 1
            self.last_clicked_row = None
            self._show_step()
        
        confirm_button = ttk.Button(
            button_frame,
            text="Confirm Header Row",
            command=confirm_header_row,
            width=20
        )
        confirm_button.pack(side=tk.LEFT, padx=5)
        
        cancel_button = ttk.Button(
            button_frame,
            text="Cancel",
            command=header_dialog.destroy,
            width=20
        )
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        # Make window fullscreen after all widgets are created
        header_dialog.update_idletasks()
        header_dialog.state('zoomed')  # Windows - maximize to fullscreen
        # For Linux/Mac, use: header_dialog.attributes('-zoomed', True)

    def _highlight_selections(self):
        """Highlight previously selected rows/columns"""
        # Clear all highlights first
        for item in self.table.get_children():
            tags = list(self.table.item(item, 'tags'))
            if 'selected_row' in tags:
                tags.remove('selected_row')
            if 'selected_col' in tags:
                tags.remove('selected_col')
            if 'current_selection' in tags:
                tags.remove('current_selection')
            self.table.item(item, tags=tuple(tags))

        # Highlight header row if selected
        if self.selected_header_row:
            for item in self.table.get_children():
                row_text = self.table.item(item, 'text')
                if row_text == str(self.selected_header_row):
                    tags = list(self.table.item(item, 'tags'))
                    tags.append('selected_row')
                    self.table.item(item, tags=tuple(tags))

        # Highlight currently selected column for current step
        step_name = self.steps[self.current_step]
        current_col = None
        
        if step_name == "depth" and self.selected_depth_col is not None:
            current_col = self.selected_depth_col
        elif step_name == "length" and self.selected_length_col is not None:
            current_col = self.selected_length_col
        elif step_name == "item_number" and self.selected_item_col is not None:
            current_col = self.selected_item_col
        elif step_name == "comments" and self.selected_comments_col is not None:
            current_col = self.selected_comments_col
        
        # Highlight current column (only the clicked cell, not entire column)
        if current_col is not None and self.last_clicked_row is not None:
            for item in self.table.get_children():
                row_text = self.table.item(item, 'text')
                if row_text == str(self.last_clicked_row):
                    tags = list(self.table.item(item, 'tags'))
                    tags.append('current_selection')
                    self.table.item(item, tags=tuple(tags))

    def _on_cell_click(self, event):
        """Handle cell click (legacy method for Treeview - now uses Canvas)"""
        # This method is kept for compatibility but should use _on_cell_click_canvas
        pass

    def _confirm_step(self):
        """Confirm current step and move to next"""
        step_name = self.steps[self.current_step]

        if step_name == "header_row":
            if self.last_clicked_row is None:
                if self.selected_header_row:
                    # Use auto-detected
                    self.last_clicked_row = self.selected_header_row
                else:
                    messagebox.showwarning("Warning", "Please click on a cell in the header row first")
                    return
            self.selected_header_row = self.last_clicked_row
            self.current_step += 1

        elif step_name == "depth":
            if self.last_clicked_col is None:
                if self.selected_depth_col is not None:
                    # Use auto-detected
                    self.last_clicked_col = self.selected_depth_col
                else:
                    messagebox.showwarning("Warning", "Please click on a cell in the depth column first")
                    return
            self.selected_depth_col = self.last_clicked_col
            self.current_step += 1

        elif step_name == "length":
            if self.last_clicked_col is None:
                if self.selected_length_col is not None:
                    # Use auto-detected
                    self.last_clicked_col = self.selected_length_col
                else:
                    messagebox.showwarning("Warning", "Please click on a cell in the length column first")
                    return
            self.selected_length_col = self.last_clicked_col
            self.current_step += 1

        elif step_name == "item_number":
            if self.last_clicked_col is None:
                if self.selected_item_col is not None:
                    # Use auto-detected
                    self.last_clicked_col = self.selected_item_col
                else:
                    messagebox.showwarning("Warning", "Please click on a cell in the item number column first")
                    return
            self.selected_item_col = self.last_clicked_col
            self.current_step += 1

        elif step_name == "comments":
            # Comments column is required
            if self.last_clicked_col is None:
                if self.selected_comments_col is not None:
                    # Use auto-detected
                    self.last_clicked_col = self.selected_comments_col
                else:
                    if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                        messagebox.showwarning(
                            "Warning", 
                            "Please click on a cell in the comments column first",
                            parent=self.dialog
                        )
                    return
            self.selected_comments_col = self.last_clicked_col
            
            # All steps completed - generate clean tally
            print("All steps completed, generating clean tally...")
            self._generate_clean_tally()
            return

        # Reset selection for next step
        self.last_clicked_row = None
        self.last_clicked_col = None

        # Show next step in the same window
        if self.current_step < len(self.steps):
            self._show_step()
        else:
            # All steps completed
            self._generate_clean_tally()

    def _generate_clean_tally(self):
        """Generate clean tally with manual column mapping"""
        print("_generate_clean_tally called")
        try:
            # Check required columns
            if self.selected_header_row is None:
                if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                    messagebox.showerror("Error", "Header row not selected", parent=self.dialog)
                return

            if self.selected_depth_col is None:
                if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                    messagebox.showerror("Error", "Depth column not selected", parent=self.dialog)
                return

            if self.selected_length_col is None:
                if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                    messagebox.showerror("Error", "Length column not selected", parent=self.dialog)
                return

            # Create column mapping and store it for use in parsing
            self.column_mapping = {
                'header_row': self.selected_header_row,
                'depth': self.selected_depth_col,
                'effective_length': self.selected_length_col,
                'item_number': self.selected_item_col if self.selected_item_col is not None else None,
                'comments': self.selected_comments_col if self.selected_comments_col is not None else None,
            }
            print(f"Column mapping created: {self.column_mapping}")

            # Delete existing cleaned and output files
            input_path = Path(self.file_path)
            
            # If file is .xls, convert to .xlsx first using pandas + xlrd
            file_to_clean = self.file_path
            if input_path.suffix.lower() == '.xls':
                import pandas as pd
                try:
                    # Read .xls file using pandas with xlrd engine
                    all_sheets = pd.read_excel(self.file_path, sheet_name=None, engine='xlrd')
                    
                    # Create .xlsx file path
                    xlsx_file = input_path.parent / f"{input_path.stem}_converted.xlsx"
                    
                    # Delete existing converted file if it exists
                    if xlsx_file.exists():
                        try:
                            xlsx_file.unlink()
                        except Exception:
                            pass
                    
                    # Write to .xlsx using openpyxl engine
                    with pd.ExcelWriter(str(xlsx_file), engine='openpyxl') as writer:
                        for sheet_name, df in all_sheets.items():
                            # Truncate sheet name if too long (Excel limit is 31 characters)
                            sheet_name_truncated = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
                            df.to_excel(writer, sheet_name=sheet_name_truncated, index=False)
                    
                    # Use converted file for cleaning
                    file_to_clean = str(xlsx_file)
                    input_path = xlsx_file
                    
                except ImportError as e:
                    messagebox.showerror(
                        "Missing Package",
                        f"xlrd package is required to convert .xls files.\n\n"
                        f"Please install it:\npip install xlrd\n\n"
                        f"Error: {str(e)}"
                    )
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to convert .xls to .xlsx:\n{str(e)}")
                    return
            
            cleaned_file = input_path.parent / f"{input_path.stem}_cleaned.xlsx"
            output_file = input_path.parent / f"{input_path.stem}_output.xlsx"

            if cleaned_file.exists():
                try:
                    os.remove(cleaned_file)
                except Exception as e:
                    print(f"Could not delete existing cleaned file: {e}")

            if output_file.exists():
                try:
                    os.remove(output_file)
                except Exception as e:
                    print(f"Could not delete existing output file: {e}")

            # Generate cleaned file (manual column mapping will be used during parsing)
            from utils.clean_excel import clean_excel
            cleaned_path = str(cleaned_file)
            
            # Show progress if dialog exists
            if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                if hasattr(self, 'status_label'):
                    self.status_label.config(text="Generating cleaned file...")
                self.dialog.update()
            
            print("Calling clean_excel...")
            clean_excel(file_to_clean, cleaned_path)
            print("clean_excel completed")
            
            # Verify cleaned file was created
            if not cleaned_file.exists():
                raise FileNotFoundError(f"Cleaned file was not created: {cleaned_file}")

            # Return result
            self.result = ('manual', cleaned_path)
            print(f"Clean tally generated: {cleaned_path}")
            print(f"Result set: {self.result}")
            
            # Clean up temporary file if it exists
            if hasattr(self, '_temp_file_to_cleanup') and self._temp_file_to_cleanup:
                try:
                    if self._temp_file_to_cleanup.exists():
                        self._temp_file_to_cleanup.unlink()
                except Exception:
                    pass
            
            # Close dialog - always close after successful generation
            print("Attempting to close dialog...")
            try:
                if hasattr(self, 'dialog'):
                    if self.dialog.winfo_exists():
                        print("Dialog exists, destroying...")
                        self.dialog.destroy()
                        print("Dialog destroyed successfully")
                    else:
                        print("Dialog already destroyed")
            except Exception as close_error:
                print(f"Error closing dialog: {close_error}")
                import traceback
                traceback.print_exc()
                # Force destroy
                try:
                    if hasattr(self, 'dialog'):
                        self.dialog.destroy()
                except:
                    pass
        except Exception as e:
            # Show error and keep dialog open
            error_msg = str(e)
            if hasattr(self, 'dialog') and self.dialog.winfo_exists():
                messagebox.showerror(
                    "Error",
                    f"Failed to generate cleaned file:\n{error_msg}",
                    parent=self.dialog
                )
                if hasattr(self, 'status_label'):
                    self.status_label.config(text=f"Error: {error_msg}", foreground="red")
            else:
                print(f"Error generating cleaned file: {error_msg}")
            import traceback
            traceback.print_exc()
            return

    def close(self):
        """Close workbook and clean up temporary files"""
        # Close workbook
        if self.workbook:
            try:
                self.workbook.close()
            except Exception:
                pass
            self.workbook = None

        # Clean up temporary file if it exists
        if hasattr(self, '_temp_file_to_cleanup') and self._temp_file_to_cleanup:
            try:
                if self._temp_file_to_cleanup.exists():
                    self._temp_file_to_cleanup.unlink()
            except Exception:
                pass

    def _cleanup_and_close(self):
        """Cleanup resources and close dialog"""
        self.close()
        if hasattr(self, 'dialog') and self.dialog.winfo_exists():
            self.dialog.destroy()


if __name__ == "__main__":
    # Test dialog
    import sys

    if len(sys.argv) < 2:
        print("Usage: python manual_column_selector.py <excel_file>")
        sys.exit(1)

    root = tk.Tk()
    root.withdraw()

    dialog = ManualColumnSelector(root, sys.argv[1])
    root.wait_window(dialog.dialog)

    if dialog.result:
        print("Result:", dialog.result)
    else:
        print("Cancelled")
