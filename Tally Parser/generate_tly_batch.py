"""
Batch script to generate TLY files from all output xlsx files
"""
import sys
import os
from pathlib import Path
from datetime import datetime

# Fix encoding for Windows console
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def generate_tly_from_output(output_xlsx, tly_file=None):
    """
    Generate TLY file from output xlsx file
    
    Args:
        output_xlsx: Path to output xlsx file
        tly_file: Path for TLY file (optional, defaults to same name with .tly extension)
    
    Returns:
        dict with results: {
            'file': filename,
            'status': 'success'/'error',
            'rows': number of rows,
            'error': error message if failed,
            'tly_file': output tly file path
        }
    """
    output_path = Path(output_xlsx)
    result = {
        'file': output_path.name,
        'status': 'error',
        'rows': 0,
        'error': None,
        'tly_file': None
    }
    
    if not output_path.exists():
        result['error'] = f"File not found: {output_xlsx}"
        return result
    
    if tly_file is None:
        tly_file = output_path.parent / f"{output_path.stem}.tly"
    else:
        tly_file = Path(tly_file)
    
    try:
        from openpyxl import load_workbook
        
        # Read Excel file
        wb = load_workbook(output_xlsx, data_only=True)
        ws = wb.active
        
        # Find column indices for "Item #" and "Depth"
        header_row = 1
        item_col = None
        depth_col = None
        
        for col_idx, cell in enumerate(ws[header_row], 1):
            cell_value = str(cell.value or "").strip()
            if "item" in cell_value.lower() and "#" in cell_value:
                item_col = col_idx
            elif "depth" in cell_value.lower():
                depth_col = col_idx
        
        if item_col is None or depth_col is None:
            result['error'] = f"Could not find 'Item #' and 'Depth' columns in the Excel file"
            return result
        
        # Extract data
        rows_data = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            item_value = ws.cell(row=row_idx, column=item_col).value
            depth_value = ws.cell(row=row_idx, column=depth_col).value
            
            # Skip empty rows
            if item_value is None and depth_value is None:
                continue
            
            # Format values
            item_str = str(item_value).strip() if item_value is not None else ""
            depth_str = str(depth_value).strip() if depth_value is not None else ""
            
            if depth_str:  # Only add rows with depth
                rows_data.append((item_str, depth_str))
        
        if not rows_data:
            result['error'] = "No data found in the Excel file"
            return result
        
        # Generate TLY file
        with open(tly_file, 'w', encoding='utf-8') as f:
            # Write header
            f.write("Run Number\tDepth of Top of Joint\n")
            
            # Write data rows
            for item, depth in rows_data:
                f.write(f"{item}\t{depth}\n")
        
        # Success
        result['status'] = 'success'
        result['rows'] = len(rows_data)
        result['tly_file'] = str(tly_file)
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


def main():
    """Main batch TLY generation function"""
    # Find all output files in test_results directory
    base_dir = Path(__file__).parent
    test_results_dir = base_dir / "test_results"
    
    if not test_results_dir.exists():
        print(f"Error: test_results directory not found: {test_results_dir}")
        return
    
    # Find all *_output.xlsx files
    output_files = list(test_results_dir.glob("*_output.xlsx"))
    
    if not output_files:
        print("No output files found in test_results directory!")
        return
    
    # Sort files by name
    output_files.sort(key=lambda x: x.name)
    
    # Run TLY generation
    print("=" * 80)
    print(f"BATCH TLY GENERATION - {len(output_files)} files")
    print("=" * 80)
    print(f"Output directory: {test_results_dir}")
    print("=" * 80)
    print()
    
    results = []
    
    for idx, output_file in enumerate(output_files, 1):
        print(f"[{idx}/{len(output_files)}] Processing: {output_file.name}")
        result = generate_tly_from_output(output_file)
        results.append(result)
        
        if result['status'] == 'success':
            print(f"  [OK] Success - {result['rows']} rows -> {Path(result['tly_file']).name}")
        else:
            print(f"  [FAIL] Failed - {result['error']}")
        print()
    
    # Generate summary report
    print("=" * 80)
    print("TLY GENERATION SUMMARY")
    print("=" * 80)
    
    success_count = sum(1 for r in results if r['status'] == 'success')
    fail_count = len(results) - success_count
    
    print(f"Total files processed: {len(results)}")
    print(f"Successful: {success_count}")
    print(f"Failed: {fail_count}")
    print()
    
    if success_count > 0:
        print("Successful files:")
        for r in results:
            if r['status'] == 'success':
                print(f"  [OK] {r['file']:<30} - {r['rows']:>4} rows -> {Path(r['tly_file']).name}")
        print()
    
    if fail_count > 0:
        print("Failed files:")
        for r in results:
            if r['status'] == 'error':
                print(f"  [FAIL] {r['file']:<30} - {r['error']}")
        print()
    
    # Save report to file
    report_file = test_results_dir / f"tly_generation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("TLY GENERATION - BATCH REPORT\n")
        f.write("=" * 80 + "\n")
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total files: {len(results)}\n")
        f.write(f"Successful: {success_count}\n")
        f.write(f"Failed: {fail_count}\n")
        f.write("=" * 80 + "\n\n")
        
        f.write("DETAILED RESULTS:\n")
        f.write("-" * 80 + "\n")
        for r in results:
            f.write(f"File: {r['file']}\n")
            f.write(f"Status: {r['status']}\n")
            if r['status'] == 'success':
                f.write(f"Rows: {r['rows']}\n")
                f.write(f"TLY File: {r['tly_file']}\n")
            else:
                f.write(f"Error: {r['error']}\n")
            f.write("-" * 80 + "\n")
    
    print(f"Report saved to: {report_file}")
    print("=" * 80)


if __name__ == "__main__":
    main()
