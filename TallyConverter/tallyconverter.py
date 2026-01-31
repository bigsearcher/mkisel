#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Tally Converter - Graphical interface for processing tally files
"""
import sys
import traceback
import warnings
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import subprocess
import platform
from pathlib import Path
import ctypes
from ctypes import wintypes

# Optional: PyInstaller splash (only available when frozen)
try:
    import pyi_splash
except ImportError:
    pyi_splash = None

# All used modules imported so PyInstaller bundles them and runtime finds them
from utils.file_converter import convert_xls_to_xlsx
from utils.clean_excel import clean_excel, clean_excel_pandas
from utils.calculate_formulas import calculate_formulas
from manual_column_selector import ManualColumnSelector
try:
    from tally_parser import parse_tally_file
except ImportError:
    try:
        from tally_parser import Parse_Tally_File as parse_tally_file
    except ImportError:
        # Define dummy to avoid NameError at runtime, though it will fail if called
        def parse_tally_file(*args, **kwargs):
            raise ImportError("Could not import parse_tally_file from tally_parser")

from excel_generator import generate_excel
from openpyxl import load_workbook


def _get_centered_geometry():
    """Fixed position; no ctypes when frozen to avoid hang on some PCs."""
    return "500x400+80+80"


class TallyConverterGUI:
    """Main GUI application for Tally Converter"""

    def __init__(self, root):
        self.root = root
        self.root.title("Tally Converter")
        # Centered or fixed position so window is never off-screen (exe on integrated graphics)
        self.root.geometry(_get_centered_geometry())
        self.root.minsize(400, 300)

        # Bind window close event to cleanup
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

        # Current file
        self.current_file = None
        self.original_file = None  # Original .xls path when opened; Manual always uses this for .xls
        self.cleaned_file = None
        self.converted_file = None  # Track converted .xls files for cleanup
        self.temp_files = []  # List of temporary files to clean up

        # Setup UI
        self._setup_ui()

        # Force window visible on some systems (integrated graphics). Safe no-op if anything fails.
        try:
            self.root.update_idletasks()
            self.root.deiconify()
            self.root.state("normal")
            self.root.lift()
            self.root.attributes("-topmost", True)
            self.root.update()
            self.root.attributes("-topmost", False)
            self.root.focus_force()
        except Exception:
            pass

        # When running as exe: repeatedly force window visible (fixes "no window" on some PCs)
        if getattr(sys, "frozen", False):
            for ms in (0, 100, 300, 800):
                self.root.after(ms, self._force_window_visible)

    def _force_window_visible(self):
        """When running as exe: reposition (center) and bring window to front."""
        try:
            self.root.geometry(_get_centered_geometry())
            self.root.deiconify()
            self.root.state("normal")
            self.root.lift()
            self.root.attributes("-topmost", True)
            self.root.update_idletasks()
            self.root.update()
            self.root.attributes("-topmost", False)
            self.root.focus_force()
        except Exception:
            pass

    def _setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Title
        title_label = ttk.Label(
            main_frame,
            text="Tally Converter",
            font=('Arial', 16, 'bold')
        )
        title_label.pack(pady=(0, 20))

        # Current file display
        self.file_label = ttk.Label(
            main_frame,
            text="No file selected",
            foreground="gray"
        )
        self.file_label.pack(pady=(0, 20))

        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.BOTH, expand=True)

        # Button 1: Open Tally File
        self.open_button = ttk.Button(
            button_frame,
            text="1. Open Tally File",
            command=self.open_file,
            width=30
        )
        self.open_button.pack(pady=10)

        # Button 2: Generate Clean Tally
        self.clean_button = ttk.Button(
            button_frame,
            text="2. Generate Tally Manually",
            command=self.generate_clean_tally,
            state=tk.DISABLED,
            width=30
        )
        self.clean_button.pack(pady=10)

        # Button 3: Generate TLY
        self.generate_button = ttk.Button(
            button_frame,
            text="3. Generate TLY",
            command=self.generate_tly,
            state=tk.DISABLED,
            width=30
        )
        self.generate_button.pack(pady=10)

        # Button Auto: Full automatic processing
        self.auto_button = ttk.Button(
            button_frame,
            text="Auto (Output + TLY)",
            command=self.auto_process,
            state=tk.DISABLED,
            width=30
        )
        self.auto_button.pack(pady=10)

        # Button 4: Readme
        self.readme_button = ttk.Button(
            button_frame,
            text="Readme",
            command=self.show_readme,
            width=30
        )
        self.readme_button.pack(pady=10)

        # Status bar
        self.status_label = ttk.Label(
            main_frame,
            text="Ready",
            relief=tk.SUNKEN,
            anchor=tk.W
        )
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X, pady=(20, 0))

    def _on_closing(self):
        """Cleanup temporary files and close application"""
        # Delete converted file if it exists
        if self.converted_file:
            try:
                converted_path = Path(self.converted_file)
                if converted_path.exists():
                    converted_path.unlink()
            except Exception:
                pass

        # Delete other temporary files
        for temp_file in self.temp_files:
            try:
                temp_path = Path(temp_file)
                if temp_path.exists():
                    temp_path.unlink()
            except Exception:
                pass

        # Close application
        self.root.destroy()

    def _open_file_in_excel(self, file_path):
        """Open file in Excel (cross-platform)"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                return
            
            system = platform.system()
            
            if system == "Windows":
                # Windows: use os.startfile or subprocess
                try:
                    os.startfile(str(file_path))
                except Exception:
                    # Fallback: try to open with excel.exe (without console window)
                    creation_flags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
                    subprocess.Popen(['excel.exe', str(file_path)], shell=False, creationflags=creation_flags)
            elif system == "Darwin":  # macOS
                subprocess.Popen(['open', '-a', 'Microsoft Excel', str(file_path)], 
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:  # Linux and others
                # Try common Linux Excel alternatives
                subprocess.Popen(['xdg-open', str(file_path)], 
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            # Silently fail - don't interrupt the workflow if Excel can't be opened
            pass

    def _open_file_in_notepad(self, file_path):
        """Open file in Notepad (Windows) or default text editor"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                return
            
            system = platform.system()
            
            if system == "Windows":
                # Windows: use notepad (without console window)
                try:
                    creation_flags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
                    subprocess.Popen(['notepad.exe', str(file_path)], 
                                   shell=False, creationflags=creation_flags)
                except Exception:
                    # Fallback: use os.startfile
                    os.startfile(str(file_path))
            elif system == "Darwin":  # macOS
                subprocess.Popen(['open', '-a', 'TextEdit', str(file_path)],
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:  # Linux and others
                # Try common Linux text editors
                subprocess.Popen(['xdg-open', str(file_path)],
                               stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            # Silently fail - don't interrupt the workflow if editor can't be opened
            pass

    def _convert_xls_to_xlsx(self, xls_file):
        """Convert .xls file to .xlsx format"""
        try:
            self.status_label.config(text="Converting .xls to .xlsx...")
            self.root.update()

            # Convert using utility function
            xlsx_file = convert_xls_to_xlsx(xls_file)

            self.status_label.config(text=f"Converted: {xlsx_file.name}")
            self.root.update()

            messagebox.showinfo(
                "File Converted",
                f"Old .xls file has been converted to .xlsx format.\n\n"
                f"Original: {Path(xls_file).name}\n"
                f"Converted: {xlsx_file.name}\n\n"
                f"You can now proceed with parsing."
            )

            return str(xlsx_file)

        except ImportError as e:
            messagebox.showerror(
                "Missing Package",
                f"xlrd package is required to convert .xls files.\n\n"
                f"Please install it:\npip install xlrd\n\n"
                f"Error: {str(e)}"
            )
            return None
        except Exception as e:
            messagebox.showerror("Conversion Error", f"Failed to convert .xls to .xlsx:\n{str(e)}")
            self.status_label.config(text="Conversion failed")
            return None

    def open_file(self):
        """Open a tally file"""
        filename = filedialog.askopenfilename(
            title="Select Tally File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xlsm"),
                ("All files", "*.*")
            ]
        )

        if filename:
            file_path = Path(filename)
            
            # Delete existing related files to avoid conflicts
            file_stem = file_path.stem
            file_dir = file_path.parent
            
            related_files = [
                file_dir / f"{file_stem}_converted.xlsx",
                file_dir / f"{file_stem}_cleaned.xlsx",
                file_dir / f"{file_stem}_output.xlsx"
            ]
            
            for related_file in related_files:
                if related_file.exists():
                    try:
                        related_file.unlink()
                    except Exception:
                        pass
            
            # Check if it's an old .xls file
            if file_path.suffix.lower() == '.xls':
                # Convert to .xlsx first
                converted_file = self._convert_xls_to_xlsx(filename)
                if not converted_file:
                    # Conversion failed, don't proceed
                    return
                # Use converted file for auto; keep original for Manual (auto may delete converted)
                self.original_file = filename
                self.current_file = converted_file
                self.converted_file = converted_file  # Track for cleanup
                original_name = file_path.name
                converted_name = Path(converted_file).name
            else:
                # Use file as-is
                self.original_file = None
                self.current_file = filename
                self.converted_file = None
                original_name = file_path.name
                converted_name = None
            
            self.cleaned_file = None

            # Update UI
            if converted_name:
                self.file_label.config(
                    text=f"Selected: {original_name} (converted to {converted_name})",
                    foreground="black"
                )
            else:
                self.file_label.config(
                    text=f"Selected: {original_name}",
                    foreground="black"
                )

            # Enable buttons
            self.clean_button.config(state=tk.NORMAL)
            # Generate TLY button can work independently (reads from any xlsx file)
            self.generate_button.config(state=tk.NORMAL)
            # Auto button enabled when file is loaded
            self.auto_button.config(state=tk.NORMAL)

            if converted_name:
                self.status_label.config(text=f"Loaded: {original_name} → {converted_name}")
            else:
                self.status_label.config(text=f"Loaded: {original_name}")

            # Automatically trigger processing
            self.root.after(100, self.auto_process)

    def _attempt_auto_process(self, input_path):
        """
        Attempt to clean and process a file automatically.
        Returns: (success, result_tuple_or_none)
        """
        try:
            # Output path
            output_path = input_path.parent / f"{input_path.stem}_cleaned.xlsx"

            # Clean the file
            try:
                clean_excel(str(input_path), str(output_path), parent_window=self.root)
                
                # Verify that cleaned file was created
                if not output_path.exists():
                    raise FileNotFoundError(f"Cleaned file was not created: {output_path}")
                
                return True, ('auto', str(output_path), None)
            except Exception as e:
                # If cleaning failed, try fallback: calculate formulas in Excel and retry
                # Fallback logic... for now let's reuse the simple fallback pattern or just fail to manual
                # Given strict instruction to just "Auto First", extensive fallback might be better here
                # replicating fallback logic from ModeSelectionDialog.use_auto_mode:
                
                print(f"Auto clean failed: {e}. Trying fallback...")
                
                # Create temporary file for Excel calculation
                temp_calculated = input_path.parent / f"{input_path.stem}_calculated_temp.xlsx"
                
                # Calculate formulas using Excel COM
                # Show progress message
                self.status_label.config(text="Opening file in Excel (Fallback)...")
                self.root.update()
                
                # Calculate formulas - this will open Excel
                calculate_formulas(str(input_path), str(temp_calculated))
                
                if not temp_calculated.exists():
                     raise FileNotFoundError(f"Calculated file was not created: {temp_calculated}")

                self.status_label.config(text="Processing calculated file...")
                self.root.update()
                
                # Try cleaning again with calculated file
                try:
                    clean_excel(str(temp_calculated), str(output_path), parent_window=self.root)
                except Exception:
                    clean_excel_pandas(str(temp_calculated), str(output_path), parent_window=self.root)
                
                if not output_path.exists():
                    raise FileNotFoundError(f"Cleaned file was not created: {output_path}")
                
                 # Clean up temporary file
                if temp_calculated.exists():
                    try:
                        temp_calculated.unlink()
                    except Exception:
                        pass
                
                return True, ('auto', str(output_path), None)

        except Exception as e:
            # Auto mode failed
            return False, e

    def _run_manual_mode(self, input_path):
        """Run manual column selector"""
        try:
            selector = ManualColumnSelector(self.root, str(input_path))
            # The dialog is modal (wait_window called inside if setup correctly or we call it here)
            # ManualColumnSelector uses grab_set, but we need to wait for it.
            # actually ManualColumnSelector shows itself. we need to wait for its window.
            self.root.wait_window(selector.dialog)

            if selector.result:
                if len(selector.result) == 3:
                     return selector.result
                else:
                     mode, cleaned_file = selector.result
                     return (mode, cleaned_file, None)
            return None
        except Exception as e:
            messagebox.showerror("Error", f"Manual mode failed:\n{str(e)}")
            return None

    def generate_clean_tally(self):
        """Generate clean tally - show Auto/Choose dialog, then automatically generate output"""
        if not self.current_file and not self.original_file:
            messagebox.showerror("Error", "Please open a file first")
            return

        # For .xls we always use the original file: auto may have deleted the converted one
        if self.original_file:
            input_path = Path(self.original_file)
            if not input_path.exists():
                messagebox.showerror("Error", f"Original file not found:\n{input_path}\n\nRe-open the file (1. Open Tally File).")
                return
            self.status_label.config(text="Converting .xls to .xlsx...")
            self.root.update()
            converted_file = self._convert_xls_to_xlsx(str(input_path))
            if not converted_file:
                return
            self.current_file = converted_file
            self.converted_file = converted_file
            input_path = Path(self.current_file)
        else:
            input_path = Path(self.current_file)
            if not input_path.exists():
                messagebox.showerror("Error", f"File not found:\n{input_path}\n\nPlease open a file first (1. Open Tally File).")
                return
            if input_path.suffix.lower() == '.xls':
                converted_file = self._convert_xls_to_xlsx(str(input_path))
                if not converted_file:
                    return
                self.current_file = converted_file
                self.converted_file = converted_file
                input_path = Path(self.current_file)

        # Delete existing cleaned and output files before generation
        cleaned_file_path = input_path.parent / f"{input_path.stem}_cleaned.xlsx"
        output_file_path = input_path.parent / f"{input_path.stem}_output.xlsx"

        if cleaned_file_path.exists():
            try:
                os.remove(cleaned_file_path)
            except Exception:
                pass

        if output_file_path.exists():
            try:
                os.remove(output_file_path)
            except Exception:
                pass

        # Show dialog: Auto or Choose column header
        
        # START MANUAL WORKFLOW
        self.status_label.config(text="Starting manual selection...")
        self.root.update()
        
        # Run manual mode directly
        result_tuple = self._run_manual_mode(input_path)
        
        column_mapping = None
        cleaned_file = None
        
        if result_tuple:
             mode, cleaned_file, column_mapping = result_tuple
        else:
            self.status_label.config(text="Manual selection cancelled")
            return

        # Check result
        if cleaned_file:
            # Proceed with processing...
            self.cleaned_file = cleaned_file
                
            # Step 1: Clean file is ready
            self.status_label.config(text=f"Cleaned file ready: {os.path.basename(cleaned_file)}")
            self.root.update()
            
            # Step 2: Automatically parse and generate output
            try:
                # Ask for output location
                output_file = filedialog.asksaveasfilename(
                    title="Save TLY Output",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile=Path(self.current_file).stem + "_output.xlsx"
                )

                if not output_file:
                    # User cancelled - just enable the button for manual generation
                    self.generate_button.config(state=tk.NORMAL)
                    self.status_label.config(text=f"Cleaned file ready: {os.path.basename(cleaned_file)}")
                    messagebox.showinfo(
                        "Success",
                        f"Clean tally generated:\n{os.path.basename(cleaned_file)}\n\nYou can generate TLY output manually using button 3."
                    )
                    return

                # Parse data
                self.status_label.config(text="Parsing data...")
                self.root.update()

                if column_mapping:
                    data = parse_tally_file(self.cleaned_file, column_mapping=column_mapping, parent_window=self.root)
                else:
                    data = parse_tally_file(self.cleaned_file, parent_window=self.root)

                if not data:
                    messagebox.showwarning("Warning", "No data extracted from file")
                    self.generate_button.config(state=tk.NORMAL)
                    return

                # Generate output
                self.status_label.config(text="Generating output file...")
                self.root.update()

                generate_excel(data, output_file)

                # Success
                self.generate_button.config(state=tk.NORMAL)
                self.status_label.config(text=f"Generated: {os.path.basename(output_file)}")
                
                # Open file in Excel
                self._open_file_in_excel(output_file)
                
                messagebox.showinfo(
                    "Success",
                    f"Clean tally and TLY output generated successfully!\n\n"
                    f"Cleaned file: {os.path.basename(cleaned_file)}\n"
                    f"Output file: {os.path.basename(output_file)}\n"
                    f"Rows processed: {len(data)}\n\n"
                    f"File opened in Excel."
                )

            except Exception as e:
                error_msg = str(e)
                # Check if it's a header detection error
                if "Could not find header row" in error_msg or "required columns" in error_msg:
                    detailed_msg = (
                        f"Could not automatically find table headers.\n\n"
                        f"Possible reasons:\n"
                        f"• Headers are not in the first 30 rows\n"
                        f"• Column names differ from expected\n"
                        f"• Table structure is non-standard\n\n"
                        f"Recommendation: Use 'Choose Column Header' mode\n"
                        f"for manual selection of column headers."
                    )
                    messagebox.showerror("Header Detection Error", detailed_msg)
                else:
                    messagebox.showerror("Error", f"Failed to generate output:\n{error_msg}")
                self.status_label.config(text="Error occurred")
                self.generate_button.config(state=tk.NORMAL)

    def auto_process(self):
        """Auto process: Generate cleaned file, output.xlsx, and TLY file automatically"""
        if not self.current_file:
            messagebox.showerror("Error", "Please open a file first")
            return

        input_path = Path(self.current_file)
        
        # If file is .xls, convert to .xlsx first
        if input_path.suffix.lower() == '.xls':
            converted_file = self._convert_xls_to_xlsx(str(input_path))
            if not converted_file:
                return
            self.current_file = converted_file
            self.converted_file = converted_file
            input_path = Path(self.current_file)

        # Delete existing files before generation
        cleaned_file_path = input_path.parent / f"{input_path.stem}_cleaned.xlsx"
        output_file_path = input_path.parent / f"{input_path.stem}_output.xlsx"
        tly_file_path = input_path.parent / f"{input_path.stem}_output.tly"

        for file_path in [cleaned_file_path, output_file_path, tly_file_path]:
            if file_path.exists():
                try:
                    os.remove(file_path)
                except Exception:
                    pass

        try:
            # Step 1: Generate cleaned file using Auto mode
            self.status_label.config(text="Generating cleaned file (Auto mode)...")
            self.root.update()

            # Use Auto mode directly (bypass dialog)
            cleaned_file = str(cleaned_file_path)
            
            try:
                clean_excel(str(input_path), cleaned_file, parent_window=self.root)
                
                # Verify that cleaned file was created
                if not Path(cleaned_file).exists():
                    raise FileNotFoundError(f"Cleaned file was not created: {cleaned_file}")
            except Exception as e:
                # If cleaning failed, try fallback: calculate formulas in Excel and retry
                try:
                    temp_calculated = input_path.parent / f"{input_path.stem}_calculated_temp.xlsx"
                    
                    self.status_label.config(text="Opening file in Excel...")
                    self.root.update()
                    
                    calculate_formulas(str(input_path), str(temp_calculated))
                    
                    if not temp_calculated.exists():
                        raise FileNotFoundError(f"Calculated file was not created: {temp_calculated}")
                    
                    self.status_label.config(text="Processing calculated file...")
                    self.root.update()
                    
                    try:
                        clean_excel(str(temp_calculated), cleaned_file, parent_window=self.root)
                    except Exception as clean_error:
                        clean_excel_pandas(str(temp_calculated), cleaned_file, parent_window=self.root)
                    
                    if not Path(cleaned_file).exists():
                        raise FileNotFoundError(f"Cleaned file was not created: {cleaned_file}")
                    
                    # Clean up temporary file
                    if temp_calculated.exists():
                        try:
                            temp_calculated.unlink()
                        except Exception:
                            pass
                except Exception as e2:
                    raise Exception(f"Failed to clean file: {str(e)}\nFallback also failed: {str(e2)}")

            # Step 2: Parse and generate output.xlsx
            self.status_label.config(text="Parsing data...")
            self.root.update()

            data = parse_tally_file(cleaned_file, parent_window=self.root)

            if not data:
                messagebox.showwarning("Warning", "No data extracted from file")
                return

            self.status_label.config(text="Generating output file...")
            self.root.update()

            output_file = str(output_file_path)
            generate_excel(data, output_file)

            # Step 3: Generate TLY file
            self.status_label.config(text="Generating TLY file...")
            self.root.update()

            wb = None
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    wb = load_workbook(output_file, data_only=True)
                ws = wb.active

                # Find column indices for "Item #" and "Depth"
                header_row = 1
                item_col = None
                depth_col = None

                for col_idx, cell in enumerate(ws[header_row], 1):
                    cell_value = str(cell.value or "").strip()
                    if "item" in cell_value.lower() and "#" in cell_value:
                        item_col = col_idx
                    elif "depth" in cell_value.lower():
                        depth_col = col_idx

                if item_col is None or depth_col is None:
                    messagebox.showerror("Error", "Could not find 'Item #' and 'Depth' columns in the output file")
                    return

                # Extract data for TLY
                rows_data = []
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    item_value = ws.cell(row=row_idx, column=item_col).value
                    depth_value = ws.cell(row=row_idx, column=depth_col).value

                    if item_value is None and depth_value is None:
                        continue

                    item_str = str(item_value).strip() if item_value is not None else ""
                    # Round depth to 2 decimal places if numeric
                    if depth_value is not None:
                        try:
                            depth_rounded = round(float(depth_value), 2)
                            depth_str = str(depth_rounded)
                        except (ValueError, TypeError):
                            depth_str = str(depth_value).strip()
                    else:
                        depth_str = ""

                    if depth_str:
                        rows_data.append((item_str, depth_str))

                if not rows_data:
                    messagebox.showwarning("Warning", "No data found for TLY file")
                    return

            finally:
                # Close workbook to free resources
                if wb:
                    try:
                        wb.close()
                    except Exception:
                        pass

            # Generate TLY file
            tly_file = str(tly_file_path)
            with open(tly_file, 'w', encoding='utf-8') as f:
                f.write("Run Number\tDepth of Top of Joint\n")
                for item, depth in rows_data:
                    f.write(f"{item}\t{depth}\n")

            # Success - open files
            self.status_label.config(text=f"Success! Generated output and TLY files")
            self.root.update()

            # Remove intermediate cleaned file only; keep converted .xlsx so Manual can use it
            if hasattr(self, 'cleaned_file') and self.cleaned_file:
                try:
                    cleaned_path = Path(self.cleaned_file)
                    if cleaned_path.exists():
                        cleaned_path.unlink()
                except Exception as e:
                    print(f"Could not delete cleaned file: {e}")

            # Do not delete converted file here — user may press "2. Generate Tally Manually" next;
            # converted file is removed in _on_closing when the app is closed.

            # Open Excel with output file
            self._open_file_in_excel(output_file)
            
            # Open Notepad with TLY file
            self._open_file_in_notepad(tly_file)
            
            messagebox.showinfo(
                "Success",
                f"Auto processing completed successfully!\n\n"
                f"Files generated:\n"
                f"• Output: {os.path.basename(output_file)}\n"
                f"• TLY: {os.path.basename(tly_file)}\n"
                f"Rows processed: {len(data)}\n\n"
                f"Files opened in Excel and Notepad."
            )

        except Exception as e:
            error_msg = str(e)
            if "Could not find header row" in error_msg or "required columns" in error_msg:
                detailed_msg = (
                    f"Could not automatically find table headers.\n\n"
                    f"Possible reasons:\n"
                    f"• Headers are not in the first 50 rows\n"
                    f"• Column names differ from expected\n"
                    f"• Table structure is non-standard\n\n"
                    f"Recommendation: Use button '2. Generate Clean Tally'\n"
                    f"with 'Choose Column Header' mode for manual selection."
                )
                messagebox.showerror("Header Detection Error", detailed_msg)
            else:
                messagebox.showerror("Error", f"Auto processing failed:\n{error_msg}")
            self.status_label.config(text="Error occurred")

    def generate_tly(self):
        """Generate final TLY output from output xlsx file"""
        try:
            # Ask user to select output xlsx file
            input_xlsx = filedialog.askopenfilename(
                title="Select Output XLSX File",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialdir=Path(self.current_file).parent if self.current_file else None
            )

            if not input_xlsx:
                return

            # Ask for TLY output location
            tly_file = filedialog.asksaveasfilename(
                title="Save TLY File",
                defaultextension=".tly",
                filetypes=[("TLY files", "*.tly"), ("Text files", "*.txt"), ("All files", "*.*")],
                initialfile=Path(input_xlsx).stem + ".tly"
            )

            if not tly_file:
                return

            # Read Excel file and extract Item # and Depth columns
            self.status_label.config(text="Reading Excel file...")
            self.root.update()

            wb = None
            try:
                wb = load_workbook(input_xlsx, data_only=True)
                ws = wb.active

                # Find column indices for "Item #" and "Depth"
                header_row = 1
                item_col = None
                depth_col = None

                for col_idx, cell in enumerate(ws[header_row], 1):
                    cell_value = str(cell.value or "").strip()
                    if "item" in cell_value.lower() and "#" in cell_value:
                        item_col = col_idx
                    elif "depth" in cell_value.lower():
                        depth_col = col_idx

                if item_col is None or depth_col is None:
                    messagebox.showerror("Error", "Could not find 'Item #' and 'Depth' columns in the Excel file")
                    return

                # Extract data
                self.status_label.config(text="Extracting data...")
                self.root.update()

                rows_data = []
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    item_value = ws.cell(row=row_idx, column=item_col).value
                    depth_value = ws.cell(row=row_idx, column=depth_col).value

                    # Skip empty rows
                    if item_value is None and depth_value is None:
                        continue

                    # Format values
                    item_str = str(item_value).strip() if item_value is not None else ""
                    depth_str = str(depth_value).strip() if depth_value is not None else ""

                    if depth_str:  # Only add rows with depth
                        rows_data.append((item_str, depth_str))

                if not rows_data:
                    messagebox.showwarning("Warning", "No data found in the Excel file")
                    return

            finally:
                # Close workbook to free resources
                if wb:
                    try:
                        wb.close()
                    except Exception:
                        pass

            # Generate TLY file
            self.status_label.config(text="Generating TLY file...")
            self.root.update()

            with open(tly_file, 'w', encoding='utf-8') as f:
                # Write header
                f.write("Run Number\tDepth of Top of Joint\n")

                # Write data rows
                for item, depth in rows_data:
                    f.write(f"{item}\t{depth}\n")

            # Success
            self.status_label.config(text=f"Generated: {os.path.basename(tly_file)}")
            
            # Open file in Notepad
            self._open_file_in_notepad(tly_file)
            
            messagebox.showinfo(
                "Success",
                f"TLY file generated successfully!\n\nRows: {len(rows_data)}\nFile: {os.path.basename(tly_file)}\n\nFile opened in Notepad."
            )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate TLY:\n{str(e)}")
            self.status_label.config(text="Error occurred")
            traceback.print_exc()

    def show_readme(self):
        """Show readme window with instructions"""
        # Create readme window
        readme_window = tk.Toplevel(self.root)
        readme_window.title("Help - Tally Converter")
        readme_window.geometry("750x650")
        readme_window.transient(self.root)
        readme_window.resizable(True, True)
        
        # Center window
        readme_window.update_idletasks()
        x = (readme_window.winfo_screenwidth() // 2) - (750 // 2)
        y = (readme_window.winfo_screenheight() // 2) - (650 // 2)
        readme_window.geometry(f"750x650+{x}+{y}")
        
        # Create main container
        main_frame = ttk.Frame(readme_window, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(
            main_frame,
            text="Tally Converter User Guide",
            font=('Arial', 14, 'bold')
        )
        title_label.pack(pady=(0, 15))
        
        # Create scrollable text widget with proper layout
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Text widget with scrollbar
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            font=('Consolas', 9),
            padx=15,
            pady=15,
            bg='white',
            relief=tk.FLAT,
            borderwidth=1
        )
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=text_widget.yview)
        
        # Readme content
        readme_text = """TALLY PARSER USER GUIDE

1. QUICK START

   • Click "1. Open Tally File" and select your Excel file.
   • Click "Auto (Output + TLY)" for automatic processing.
   • If automatic processing fails, try "2. Generate Tally Manually".
   
   That's it!

2. IF AUTOMATIC PROCESSING FAILS

   If "Auto" mode fails or cannot find headers:
   
   • Click "2. Generate Tally Manually".
   • A window will appear where you can select the columns manually:
     1. Click on the ROW that contains the headers (Depth, Length, etc.).
     2. Click on the COLUMN for "Depth".
     3. Click on the COLUMN for "Length".
     4. Click on the COLUMN for "Item Number".
     5. Click on the COLUMN for "Comments".
   • Once finished, the program will process the file using your selection.

3. BUTTONS EXPLAINED

   [1. Open Tally File]
   Opens a dialog to select your Excel file (.xls or .xlsx).
   Old .xls files are automatically converted to .xlsx.

   [2. Generate Tally Manually]
   Manual processing mode (or retry mode).
   • Allows you to manually select the header row and required columns.
   • Useful if the "Auto" button fails to detect the table structure.
   • Creates two files: 
     - "_output.xlsx" (Clean Excel table)
     - ".tly" (Text file for Tally)

   [3. Generate TLY]
   Useful if you already have a clean "output.xlsx" file and just want to 
   regenerate the .tly file without parsing everything again.

   [Auto (Output + TLY)]
   The recommended first step. Automatically attempts to:
   1. Clean the file (remove images, fix formulas).
   2. Detect headers and data columns.
   3. Generate both Output and TLY files.

   [Readme]
   Shows this help window.

4. TROUBLESHOOTING

   • "Macros" / "Formulas" errors: The program tries to fix these automatically 
     by opening Excel in the background. If this fails, try opening the file 
     in Excel yourself, saving it, and trying again.
   • "Header not found": Use "2. Generate Tally Manually".
"""
        
        text_widget.insert(tk.END, readme_text)
        text_widget.config(state=tk.DISABLED)  # Make read-only
        
        # Close button
        close_button = ttk.Button(
            main_frame,
            text="Close",
            command=readme_window.destroy
        )
        close_button.pack(pady=(10, 0))





def _try_activate_existing_instance():
    """
    On Windows: if another instance is already running and responding, bring its window to front and return True.
    If the existing window is hung, ask user whether to start a new instance. Returns True to exit, False to start.
    """
    if platform.system() != "Windows":
        return False
    try:
        user32 = ctypes.windll.user32

        target_title = "Tally Converter"
        found = [None]

        def enum_callback(hwnd, _):
            length = user32.GetWindowTextLengthW(hwnd) + 1
            if length <= 1:
                return True
            buf = ctypes.create_unicode_buffer(length)
            user32.GetWindowTextW(hwnd, buf, length)
            if buf.value.strip() == target_title:
                found[0] = hwnd
                return False  # stop enumeration
            return True

        WNDENUMPROC = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
        user32.EnumWindows(WNDENUMPROC(enum_callback), 0)
        if not found[0]:
            return False  # no existing window, start normally

        hwnd = found[0]
        # If existing window is hung, let user start a new instance
        if user32.IsHungAppWindow(hwnd):
            # MB_YESNO=4, MB_ICONQUESTION=0x20, IDYES=6
            r = user32.MessageBoxW(
                None,
                "The previous instance is not responding.\nStart a new instance?",
                "Tally Converter",
                4 | 0x20,
            )
            if r == 6:  # IDYES
                return False  # start new instance
            return True  # user chose No, exit

        user32.ShowWindow(hwnd, 9)  # SW_RESTORE
        user32.SetForegroundWindow(hwnd)
        return True  # activated existing window, exit
    except Exception:
        return False


def main():
    """Main entry point"""
    def log(msg):
        print(msg, flush=True)

    try:
        if getattr(sys, "frozen", False):
            log("Tally Converter starting...")
        if _try_activate_existing_instance():
            if getattr(sys, "frozen", False):
                log("Another instance is running. Exiting.")
            return
        if getattr(sys, "frozen", False):
            log("Creating window...")
        root = tk.Tk()
        # Show "Loading..." immediately on slow PCs (process grows 4→9 MB before main window)
        root.withdraw()
        splash = tk.Toplevel(root)
        splash.title("Tally Converter")
        splash.geometry("340x110")
        splash.resizable(False, False)
        ttk.Label(splash, text="Loading...", font=("Segoe UI", 11)).pack(pady=(20, 2))
        ttk.Label(splash, text="First run may take 1–2 min.", font=("Segoe UI", 9), foreground="gray").pack(pady=(0, 20))
        splash.update_idletasks()
        splash.update()
        app = TallyConverterGUI(root)
        try:
            splash.destroy()
        except Exception:
            pass
        root.deiconify()
        # Close bootloader splash (shown before Python started) so only main window remains
        if getattr(sys, "frozen", False) and pyi_splash is not None:
            try:
                pyi_splash.close()
            except Exception:
                pass
        if getattr(sys, "frozen", False):
            log("Window created. Starting main loop (close window to exit).")
        # When exe: force window visible once more before event loop (helps on some PCs)
        if getattr(sys, "frozen", False):
            try:
                root.update_idletasks()
                root.update()
                root.deiconify()
                root.lift()
                root.attributes("-topmost", True)
                root.update()
                root.attributes("-topmost", False)
                root.focus_force()
            except Exception:
                pass
        root.mainloop()
        if getattr(sys, "frozen", False):
            log("Exiting.")
    except Exception as e:
        log(f"Error: {e}")
        traceback.print_exc()
        if getattr(sys, "frozen", False):
            input("Press Enter to close...")


if __name__ == "__main__":
    main()
